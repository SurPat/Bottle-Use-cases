
'''
Description : Machine Vision
Author      : LTTS-Machine Vision Team
Edited on   : 20-Aug-2020
Version     : V2.1

'''

from __future__ import print_function
import os
import sys
import cv2
import math
import time
import xlwt
import imutils
import sqlite3
import datetime
import argparse
import numpy as np
import pandas as pd
from glob import glob
from PyQt5.QtGui import *
from sqlite3 import Error
from pyzbar import pyzbar
from openpyxl import Workbook
from PyQt5.QtGui import QIcon
from matplotlib import pyplot as plt
from imutils.video import VideoStream
from skimage.measure import compare_ssim
from imutils import perspective,contours
from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtCore import QCoreApplication
from scipy.spatial import distance as dist
from PyQt5.QtWidgets import QApplication, QWidget, QMainWindow, QLineEdit,QAction, QMessageBox,QPushButton,QSlider,QLabel,QDialog,QApplication,QHBoxLayout,QFileDialog,QInputDialog
from openpyxl.styles import NamedStyle, Font, Border, Side,Alignment,PatternFill,Side,Protection,GradientFill

class Ui_second_Window(QtWidgets.QWidget):
    """ Second Window for Color Selection """
    submitted = QtCore.pyqtSignal(str)

    def __init__(self):
        super().__init__()
        self.resize(100, 140)
        self.title = "Color Selector"
        self.setWindowTitle(self.title)
        self.setWindowIcon(QtGui.QIcon('Color-picker-icon.png'))
        self.Red_Btn = QtWidgets.QPushButton()
        self.Green_Btn = QtWidgets.QPushButton()
        self.Yellow_Btn = QtWidgets.QPushButton()
        self.Blue_Btn = QtWidgets.QPushButton()
        self.Black_Btn = QtWidgets.QPushButton()
        self.White_Btn = QtWidgets.QPushButton()
        self.Green_Btn.setStyleSheet("font-size:35px;\n"
                                     "border-width: 2px;\n"
                                     "border-radius: 15px;\n"
                                     "border-color: black;\n"
                                     "background-color:Green;\n"
                                     "padding : 4px;")

        self.Red_Btn.setStyleSheet("font-size:35px;\n"
                                   "border-style: outset;\n"
                                   "border-width: 2px;\n"
                                   "border-radius: 15px;\n"
                                   "border-color: black;\n"
                                   "background-color:Red;\n"
                                   "padding : 4px;")

        self.Yellow_Btn.setStyleSheet("font-size:35px;\n"
                                      "border-style: outset;\n"
                                      "border-width: 2px;\n"
                                      "border-radius: 15px;\n"
                                      "border-color: black;\n"
                                      "background-color:Yellow;\n"
                                      "padding : 4px;")

        self.Blue_Btn.setStyleSheet("font-size:35px;\n"
                                    "border-style: outset;\n"
                                    "border-width: 2px;\n"
                                    "border-radius: 15px;\n"
                                    "border-color: black;\n"
                                    "background-color:Blue;\n"
                                    "padding : 4px;")

        self.Black_Btn.setStyleSheet("font-size:35px;\n"
                                     "border-style: outset;\n"
                                     "border-width: 2px;\n"
                                     "border-radius: 15px;\n"
                                     "border-color: black;\n"
                                     "background-color:Black;\n"
                                     "padding : 4px;")

        self.White_Btn.setStyleSheet("font-size:35px;\n"
                                     "border-style: outset;\n"
                                     "border-width: 2px;\n"
                                     "border-radius: 15px;\n"
                                     "border-color: black;\n"
                                     "background-color:White;\n"
                                     "padding : 4px;")

        self.setLayout(QtWidgets.QFormLayout())

        buttons = QtWidgets.QWidget()
        buttons1 = QtWidgets.QWidget()
        buttons1.setLayout(QtWidgets.QHBoxLayout())
        buttons.setLayout(QtWidgets.QHBoxLayout())
        buttons.layout().addWidget(self.Red_Btn)
        buttons.layout().addWidget(self.Green_Btn)
        buttons.layout().addWidget(self.Yellow_Btn)
        buttons1.layout().addWidget(self.Blue_Btn)
        buttons1.layout().addWidget(self.Black_Btn)
        buttons1.layout().addWidget(self.White_Btn)
        self.layout().addRow('', buttons)
        self.layout().addRow('', buttons1)

        # button Press Event to initialize main window
        self.Red_Btn.clicked.connect(self.Red_Identification)
        self.Yellow_Btn.clicked.connect(self.Yellow_Identifiction)
        self.Blue_Btn.clicked.connect(self.Blue_Identification)
        self.Black_Btn.clicked.connect(self.Black_Identification)
        self.White_Btn.clicked.connect(self.White_Identification)
        self.Green_Btn.clicked.connect(self.Green_Identification)

    # Red button Event
    def Red_Identification(self):
        Select_Color = "red"
        self.submitted.emit(Select_Color)
        self.close()

    # Yellow button Event
    def Yellow_Identifiction(self):
        Select_Color = "Yellow"
        self.submitted.emit(Select_Color)
        self.close()

    # Blue button Event
    def Blue_Identification(self):
        Select_Color = "Blue"
        self.submitted.emit(Select_Color)
        self.close()

    # Black button Event
    def Black_Identification(self):
        Select_Color = "Black"
        self.submitted.emit(Select_Color)
        self.closeEvent()

    # White button Event
    def White_Identification(self):
        Select_Color = "White"
        self.submitted.emit(Select_Color)
        self.Ui_second_Window.Close()

    # Green button Event
    def Green_Identification(self):
        Select_Color = "Green"
        self.submitted.emit(Select_Color)
        #self.Ui_second_Window.hide()

class Ui_MainWindow(QtWidgets.QWidget):

    def setupUi(self, MainWindow):
        MainWindow.setObjectName("MainWindow")
        MainWindow.resize(992, 568)
        MainWindow.setCursor(QtGui.QCursor(QtCore.Qt.IBeamCursor))
        MainWindow.setStyleSheet("\n"
"alternate-background-color: rgb(212, 225, 255);")
        self.centralwidget = QtWidgets.QWidget(MainWindow)
        self.centralwidget.setObjectName("centralwidget")
        self.OutPutScreen = QtWidgets.QLabel(self.centralwidget)
        self.OutPutScreen.setGeometry(QtCore.QRect(260, 70, 511, 361))
        self.OutPutScreen.setFrameShape(QtWidgets.QFrame.Box)
        self.OutPutScreen.setFrameShadow(QtWidgets.QFrame.Raised)
        self.OutPutScreen.setLineWidth(6)
        self.OutPutScreen.setText("")
        self.OutPutScreen.setObjectName("OutPutScreen")
        self.Start = QtWidgets.QPushButton(self.centralwidget)
        self.Start.setGeometry(QtCore.QRect(10, 310, 61, 21))
        font = QtGui.QFont()
        font.setPointSize(-1)
        self.Start.setFont(font)
        self.Start.setCursor(QtGui.QCursor(QtCore.Qt.UpArrowCursor))
        self.Start.setTabletTracking(False)
        self.Start.setAutoFillBackground(False)
        self.Start.setStyleSheet("font-size:13px;\n"
"border-style: outset;\n"
"border-width: 1px;\n"
"border-radius: 10px;\n"
"border-color: white;\n"
"background-color:green;\n"
"padding : 2px;")
        self.Start.setObjectName("Start")
        self.textBrowser = QtWidgets.QTextBrowser(self.centralwidget)
        self.textBrowser.setGeometry(QtCore.QRect(10, 120, 61, 21))
        self.textBrowser.setFrameShape(QtWidgets.QFrame.NoFrame)
        self.textBrowser.setFrameShadow(QtWidgets.QFrame.Plain)
        self.textBrowser.setLineWidth(1)
        self.textBrowser.setObjectName("textBrowser")
        self.Threshold_Value = QtWidgets.QLabel(self.centralwidget)
        self.Threshold_Value.setGeometry(QtCore.QRect(20, 150, 41, 20))
        font = QtGui.QFont()
        font.setPointSize(8)
        font.setBold(True)
        font.setWeight(75)
        self.Threshold_Value.setFont(font)
        self.Threshold_Value.setFrameShape(QtWidgets.QFrame.Box)
        self.Threshold_Value.setFrameShadow(QtWidgets.QFrame.Raised)
        self.Threshold_Value.setLineWidth(2)
        self.Threshold_Value.setText("")
        self.Threshold_Value.setObjectName("Threshold_Value")
        self.InspectionOptions = QtWidgets.QComboBox(self.centralwidget)
        self.InspectionOptions.setGeometry(QtCore.QRect(10, 70, 161, 41))
        self.InspectionOptions.setObjectName("InspectionOptions")
        self.InspectionOptions.addItem("")
        self.InspectionOptions.addItem("")
        self.InspectionOptions.addItem("")
        self.InspectionOptions.addItem("")
        self.InspectionOptions.addItem("")
        self.InspectionOptions.addItem("")
        self.InspectionOptions.addItem("")
        self.InspectionOptions.addItem("")
        self.InspectionOptions.addItem("")
        self.InspectionOptions.addItem("")
        self.InspectionOptions.addItem("")
        self.InspectionOptions.addItem("")
        self.Input_Text_2 = QtWidgets.QTextBrowser(self.centralwidget)
        self.Input_Text_2.setGeometry(QtCore.QRect(450, 30, 141, 31))
        self.Input_Text_2.setStyleSheet("font: 75 8pt \"MS Shell Dlg 2\";")
        self.Input_Text_2.setFrameShape(QtWidgets.QFrame.NoFrame)
        self.Input_Text_2.setFrameShadow(QtWidgets.QFrame.Plain)
        self.Input_Text_2.setLineWidth(1)
        self.Input_Text_2.setObjectName("Input_Text_2")
        self.textBrowser_2 = QtWidgets.QTextBrowser(self.centralwidget)
        self.textBrowser_2.setGeometry(QtCore.QRect(90, 120, 61, 21))
        self.textBrowser_2.setFrameShape(QtWidgets.QFrame.NoFrame)
        self.textBrowser_2.setFrameShadow(QtWidgets.QFrame.Plain)
        self.textBrowser_2.setLineWidth(1)
        self.textBrowser_2.setObjectName("textBrowser_2")
        self.Iterations_slider = QtWidgets.QTextBrowser(self.centralwidget)
        self.Iterations_slider.setGeometry(QtCore.QRect(170, 120, 61, 21))
        self.Iterations_slider.setFrameShape(QtWidgets.QFrame.NoFrame)
        self.Iterations_slider.setFrameShadow(QtWidgets.QFrame.Plain)
        self.Iterations_slider.setLineWidth(1)
        self.Iterations_slider.setObjectName("Iterations_slider")
        self.progressBar = QtWidgets.QProgressBar(self.centralwidget)
        self.progressBar.setGeometry(QtCore.QRect(10, 280, 191, 16))
        self.progressBar.setProperty("value", 0)
        self.progressBar.setObjectName("progressBar")
        self.ExitScreen = QtWidgets.QPushButton(self.centralwidget)
        self.ExitScreen.setGeometry(QtCore.QRect(170, 310, 61, 21))
        self.ExitScreen.setStyleSheet("font-size:13px;\n"
"border-style: outset;\n"
"border-width: 1px;\n"
"border-radius: 10px;\n"
"border-color: white;\n"
"background-color:Red;\n"
"padding : 2px;")
        self.ExitScreen.setObjectName("ExitScreen")
        self.Thrash_slider = QtWidgets.QSlider(self.centralwidget)
        self.Thrash_slider.setGeometry(QtCore.QRect(30, 170, 20, 101))
        self.Thrash_slider.setMaximum(255)
        self.Thrash_slider.setSliderPosition(127)
        self.Thrash_slider.setOrientation(QtCore.Qt.Vertical)
        self.Thrash_slider.setInvertedAppearance(True)
        self.Thrash_slider.setTickPosition(QtWidgets.QSlider.TicksAbove)
        self.Thrash_slider.setObjectName("Thrash_slider")
        self.Kernel_slider = QtWidgets.QSlider(self.centralwidget)
        self.Kernel_slider.setGeometry(QtCore.QRect(110, 170, 20, 101))
        self.Kernel_slider.setMaximum(20)
        self.Kernel_slider.setOrientation(QtCore.Qt.Vertical)
        self.Kernel_slider.setInvertedAppearance(True)
        self.Kernel_slider.setTickPosition(QtWidgets.QSlider.TicksAbove)
        self.Kernel_slider.setObjectName("Kernel_slider")
        self.Kernel_value = QtWidgets.QLabel(self.centralwidget)
        self.Kernel_value.setGeometry(QtCore.QRect(100, 150, 41, 21))
        font = QtGui.QFont()
        font.setPointSize(8)
        font.setBold(True)
        font.setWeight(75)
        self.Kernel_value.setFont(font)
        self.Kernel_value.setFrameShape(QtWidgets.QFrame.Box)
        self.Kernel_value.setFrameShadow(QtWidgets.QFrame.Raised)
        self.Kernel_value.setLineWidth(2)
        self.Kernel_value.setText("")
        self.Kernel_value.setObjectName("Kernel_value")
        self.horizontalSlider_3 = QtWidgets.QSlider(self.centralwidget)
        self.horizontalSlider_3.setGeometry(QtCore.QRect(190, 170, 20, 101))
        self.horizontalSlider_3.setMaximum(20)
        self.horizontalSlider_3.setOrientation(QtCore.Qt.Vertical)
        self.horizontalSlider_3.setInvertedAppearance(True)
        self.horizontalSlider_3.setTickPosition(QtWidgets.QSlider.TicksAbove)
        self.horizontalSlider_3.setObjectName("horizontalSlider_3")
        self.Iteration_Value = QtWidgets.QLabel(self.centralwidget)
        self.Iteration_Value.setGeometry(QtCore.QRect(180, 150, 41, 21))
        font = QtGui.QFont()
        font.setPointSize(8)
        font.setBold(True)
        font.setWeight(75)
        self.Iteration_Value.setFont(font)
        self.Iteration_Value.setFrameShape(QtWidgets.QFrame.Box)
        self.Iteration_Value.setFrameShadow(QtWidgets.QFrame.Raised)
        self.Iteration_Value.setLineWidth(2)
        self.Iteration_Value.setText("")
        self.Iteration_Value.setObjectName("Iteration_Value")
        self.label = QtWidgets.QLabel(self.centralwidget)
        self.label.setGeometry(QtCore.QRect(-20, 0, 311, 41))
        self.label.setText("")
        self.label.setPixmap(QtGui.QPixmap("lnt_logodownload.jpg"))
        self.label.setObjectName("label")
        self.Reset = QtWidgets.QPushButton(self.centralwidget)
        self.Reset.setGeometry(QtCore.QRect(90, 310, 61, 21))
        font = QtGui.QFont()
        font.setPointSize(-1)
        self.Reset.setFont(font)
        self.Reset.setCursor(QtGui.QCursor(QtCore.Qt.UpArrowCursor))
        self.Reset.setTabletTracking(False)
        self.Reset.setAutoFillBackground(False)
        self.Reset.setStyleSheet("font-size:13px;\n"
"border-style: outset;\n"
"border-width: 1px;\n"
"border-radius: 10px;\n"
"border-color: white;\n"
"background-color:Orange;\n"
"padding : 2px;")
        self.Reset.setObjectName("Reset")
        self.New_Master_Image = QtWidgets.QPushButton(self.centralwidget)
        self.New_Master_Image.setGeometry(QtCore.QRect(10, 340, 131, 31))
        font = QtGui.QFont()
        font.setPointSize(-1)
        self.New_Master_Image.setFont(font)
        self.New_Master_Image.setCursor(QtGui.QCursor(QtCore.Qt.UpArrowCursor))
        self.New_Master_Image.setTabletTracking(False)
        self.New_Master_Image.setAutoFillBackground(False)
        self.New_Master_Image.setStyleSheet("font-size:12px;\n"
"border-style: outset;\n"
"border-width: 1px;\n"
"border-radius: 10px;\n"
"border-color: white;\n"
"background-color:aqua;\n"
"padding : 2px;")
        self.New_Master_Image.setObjectName("New_Master_Image")
        self.lineEdit = QtWidgets.QLineEdit(self.centralwidget)
        self.lineEdit.setGeometry(QtCore.QRect(170, 340, 31, 21))
        self.lineEdit.setObjectName("lineEdit")
        self.customize_Image = QtWidgets.QPushButton(self.centralwidget)
        self.customize_Image.setGeometry(QtCore.QRect(10, 380, 131, 31))
        font = QtGui.QFont()
        font.setPointSize(-1)
        self.customize_Image.setFont(font)
        self.customize_Image.setCursor(QtGui.QCursor(QtCore.Qt.UpArrowCursor))
        self.customize_Image.setTabletTracking(False)
        self.customize_Image.setAutoFillBackground(False)
        self.customize_Image.setStyleSheet("font-size:12px;\n"
"border-style: outset;\n"
"border-width: 1px;\n"
"border-radius: 10px;\n"
"border-color: white;\n"
"background-color:aqua;\n"
"padding : 2px;")
        self.customize_Image.setObjectName("customize_Image")
        self.OutPutData = QtWidgets.QLabel(self.centralwidget)
        self.OutPutData.setGeometry(QtCore.QRect(260, 470, 511, 41))
        self.OutPutData.setFrameShape(QtWidgets.QFrame.StyledPanel)
        self.OutPutData.setFrameShadow(QtWidgets.QFrame.Plain)
        self.OutPutData.setLineWidth(6)
        self.OutPutData.setText("")
        self.OutPutData.setObjectName("OutPutData")
        self.verify_Excel = QtWidgets.QPushButton(self.centralwidget)
        self.verify_Excel.setGeometry(QtCore.QRect(100, 490, 91, 21))
        font = QtGui.QFont()
        font.setPointSize(-1)
        self.verify_Excel.setFont(font)
        self.verify_Excel.setStyleSheet("font-size:12px;\n"
"border-style: outset;\n"
"border-width: 1px;\n"
"border-radius: 10px;\n"
"border-color: white;\n"
"background-color:Grey;\n"
"padding : 2px;")
        self.verify_Excel.setObjectName("verify_Excel")
        self.Help = QtWidgets.QPushButton(self.centralwidget)
        self.Help.setGeometry(QtCore.QRect(10, 490, 71, 21))
        font = QtGui.QFont()
        font.setPointSize(-1)
        self.Help.setFont(font)
        self.Help.setStyleSheet("font-size:12px;\n"
"border-style: outset;\n"
"border-width: 1px;\n"
"border-radius: 10px;\n"
"border-color: white;\n"
"background-color:Grey;\n"
"padding : 2px;")
        self.Help.setObjectName("Help")
        self.AllOutPutData = QtWidgets.QLabel(self.centralwidget)
        self.AllOutPutData.setGeometry(QtCore.QRect(790, 70, 171, 441))
        self.AllOutPutData.setFrameShape(QtWidgets.QFrame.StyledPanel)
        self.AllOutPutData.setFrameShadow(QtWidgets.QFrame.Plain)
        self.AllOutPutData.setLineWidth(6)
        self.AllOutPutData.setText("")
        self.AllOutPutData.setObjectName("AllOutPutData")
        self.Input_Text_3 = QtWidgets.QTextBrowser(self.centralwidget)
        self.Input_Text_3.setGeometry(QtCore.QRect(810, 30, 131, 31))
        self.Input_Text_3.setStyleSheet("font: 75 8pt \"MS Shell Dlg 2\";")
        self.Input_Text_3.setFrameShape(QtWidgets.QFrame.NoFrame)
        self.Input_Text_3.setFrameShadow(QtWidgets.QFrame.Plain)
        self.Input_Text_3.setLineWidth(1)
        self.Input_Text_3.setObjectName("Input_Text_3")
        self.Input_Text_4 = QtWidgets.QTextBrowser(self.centralwidget)
        self.Input_Text_4.setGeometry(QtCore.QRect(260, 440, 131, 21))
        self.Input_Text_4.setStyleSheet("font: 75 8pt \"MS Shell Dlg 2\";")
        self.Input_Text_4.setFrameShape(QtWidgets.QFrame.NoFrame)
        self.Input_Text_4.setFrameShadow(QtWidgets.QFrame.Plain)
        self.Input_Text_4.setLineWidth(1)
        self.Input_Text_4.setObjectName("Input_Text_4")
        self.frame = QtWidgets.QFrame(self.centralwidget)
        self.frame.setGeometry(QtCore.QRect(160, 400, 120, 80))
        self.frame.setFrameShape(QtWidgets.QFrame.StyledPanel)
        self.frame.setFrameShadow(QtWidgets.QFrame.Raised)
        self.frame.setObjectName("frame")
        MainWindow.setCentralWidget(self.centralwidget)
        self.statusbar = QtWidgets.QStatusBar(MainWindow)
        self.statusbar.setObjectName("statusbar")
        MainWindow.setStatusBar(self.statusbar)
        self.menuBar = QtWidgets.QMenuBar(MainWindow)
        self.menuBar.setGeometry(QtCore.QRect(0, 0, 992, 21))
        self.menuBar.setObjectName("menuBar")
        self.menuFile = QtWidgets.QMenu(self.menuBar)
        self.menuFile.setObjectName("menuFile")
        self.menuEdit = QtWidgets.QMenu(self.menuBar)
        self.menuEdit.setObjectName("menuEdit")
        self.menuView = QtWidgets.QMenu(self.menuBar)
        self.menuView.setObjectName("menuView")
        self.menuSetting = QtWidgets.QMenu(self.menuBar)
        self.menuSetting.setObjectName("menuSetting")
        self.menuHelp = QtWidgets.QMenu(self.menuBar)
        self.menuHelp.setObjectName("menuHelp")
        MainWindow.setMenuBar(self.menuBar)
        self.toolBar = QtWidgets.QToolBar(MainWindow)
        self.toolBar.setObjectName("toolBar")
        MainWindow.addToolBar(QtCore.Qt.TopToolBarArea, self.toolBar)
        self.actionsave = QtWidgets.QAction(MainWindow)
        self.actionsave.setObjectName("actionsave")
        self.actionsave222 = QtWidgets.QAction(MainWindow)
        self.actionsave222.setObjectName("actionsave222")
        self.actionccccc = QtWidgets.QAction(MainWindow)
        self.actionccccc.setObjectName("actionccccc")
        self.actionhhh = QtWidgets.QAction(MainWindow)
        self.actionhhh.setObjectName("actionhhh")
        self.actionRedo = QtWidgets.QAction(MainWindow)
        self.actionRedo.setObjectName("actionRedo")
        self.actionCut = QtWidgets.QAction(MainWindow)
        self.actionCut.setObjectName("actionCut")
        self.actionPreference = QtWidgets.QAction(MainWindow)
        self.actionPreference.setObjectName("actionPreference")
        self.actionAction_Editor = QtWidgets.QAction(MainWindow)
        self.actionAction_Editor.setObjectName("actionAction_Editor")
        self.actionToolbar = QtWidgets.QAction(MainWindow)
        self.actionToolbar.setObjectName("actionToolbar")
        self.actionQT_GUI_Help = QtWidgets.QAction(MainWindow)
        self.actionQT_GUI_Help.setObjectName("actionQT_GUI_Help")
        self.actionRoom_Address_Sector_20_Welcome_sweets_junction_Atharava_gym_opposite_road_Matru_Pitru_Chaya_Apt_New_Apartment_Flat_No_105_Airoli_Mumbai_400708 = QtWidgets.QAction(MainWindow)
        self.actionRoom_Address_Sector_20_Welcome_sweets_junction_Atharava_gym_opposite_road_Matru_Pitru_Chaya_Apt_New_Apartment_Flat_No_105_Airoli_Mumbai_400708.setObjectName("actionRoom_Address_Sector_20_Welcome_sweets_junction_Atharava_gym_opposite_road_Matru_Pitru_Chaya_Apt_New_Apartment_Flat_No_105_Airoli_Mumbai_400708")
        self.menuFile.addSeparator()
        self.menuFile.addSeparator()
        self.menuFile.addAction(self.actionsave)
        self.menuFile.addAction(self.actionsave222)
        self.menuFile.addSeparator()
        self.menuFile.addAction(self.actionccccc)
        self.menuFile.addAction(self.actionhhh)
        self.menuEdit.addAction(self.actionRedo)
        self.menuEdit.addAction(self.actionCut)
        self.menuView.addAction(self.actionAction_Editor)
        self.menuView.addAction(self.actionToolbar)
        self.menuSetting.addAction(self.actionPreference)
        self.menuBar.addAction(self.menuFile.menuAction())
        self.menuBar.addAction(self.menuEdit.menuAction())
        self.menuBar.addAction(self.menuView.menuAction())
        self.menuBar.addAction(self.menuSetting.menuAction())
        self.menuBar.addAction(self.menuHelp.menuAction())
        
        # ----- SKC ----- #
        self.customize_Image.clicked.connect(self.customize_Master_Image)  # New_Master_Image
        self.aa = True
        self.label = QtWidgets.QLabel(self.centralwidget)
        self.label.setGeometry(QtCore.QRect(50, 550, 161, 41))
        self.label.setObjectName("label")
        MainWindow.setCentralWidget(self.centralwidget)
        self.Start.clicked.connect(self.pressed)
        self.ExitScreen.clicked.connect(self.Exit_Screen)
        self.Reset.clicked.connect(self.ResetOption)
        self.New_Master_Image.clicked.connect(self.MasterImageRetrieved)
        self.verify_Excel.clicked.connect(self.verifyExcel) # self.verify_Excel
        # ----- SKC ----- #

        self.retranslateUi(MainWindow)
        QtCore.QMetaObject.connectSlotsByName(MainWindow)

    def retranslateUi(self, MainWindow):
        _translate = QtCore.QCoreApplication.translate
        MainWindow.setWindowTitle(_translate("MainWindow", "MachineVision-A Gate Way for Industrial Application"))
        self.Start.setText(_translate("MainWindow", "Start"))
        self.textBrowser.setHtml(_translate("MainWindow", "<!DOCTYPE HTML PUBLIC \"-//W3C//DTD HTML 4.0//EN\" \"http://www.w3.org/TR/REC-html40/strict.dtd\">\n"
"<html><head><meta name=\"qrichtext\" content=\"1\" /><style type=\"text/css\">\n"
"p, li { white-space: pre-wrap; }\n"
"</style></head><body style=\" font-family:\'MS Shell Dlg 2\'; font-size:8.25pt; font-weight:400; font-style:normal;\">\n"
"<p align=\"center\" style=\" margin-top:0px; margin-bottom:0px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><span style=\" font-family:\'MS Shell Dlg 2\'; font-size:6pt; font-weight:600;\">Threshold</span></p></body></html>"))
        self.InspectionOptions.setItemText(0, _translate("MainWindow", "Inspection Options"))
        self.InspectionOptions.setItemText(1, _translate("MainWindow", "Scan QR/Barcode"))
        self.InspectionOptions.setItemText(2, _translate("MainWindow", "Colour Identification"))
        self.InspectionOptions.setItemText(3, _translate("MainWindow", "Multiple Color Identification"))
        self.InspectionOptions.setItemText(4, _translate("MainWindow", "Difference"))
        self.InspectionOptions.setItemText(5, _translate("MainWindow", "Edge Detection"))
        self.InspectionOptions.setItemText(6, _translate("MainWindow", "Object Dimension"))
        self.InspectionOptions.setItemText(7, _translate("MainWindow", "Sorting"))
        self.InspectionOptions.setItemText(8, _translate("MainWindow", "Colour Concentration"))
        self.InspectionOptions.setItemText(9, _translate("MainWindow", "Seal Check"))
        self.InspectionOptions.setItemText(10, _translate("MainWindow", "Liquid Level Inspection"))
        self.InspectionOptions.setItemText(11, _translate("MainWindow", "Almond Quality"))
        self.Input_Text_2.setHtml(_translate("MainWindow", "<!DOCTYPE HTML PUBLIC \"-//W3C//DTD HTML 4.0//EN\" \"http://www.w3.org/TR/REC-html40/strict.dtd\">\n"
"<html><head><meta name=\"qrichtext\" content=\"1\" /><style type=\"text/css\">\n"
"p, li { white-space: pre-wrap; }\n"
"</style></head><body style=\" font-family:\'MS Shell Dlg 2\'; font-size:8pt; font-weight:72; font-style:normal;\">\n"
"<p align=\"center\" style=\" margin-top:0px; margin-bottom:0px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><span style=\" font-size:9pt; font-weight:600;\">Processing-Image</span></p></body></html>"))
        self.textBrowser_2.setHtml(_translate("MainWindow", "<!DOCTYPE HTML PUBLIC \"-//W3C//DTD HTML 4.0//EN\" \"http://www.w3.org/TR/REC-html40/strict.dtd\">\n"
"<html><head><meta name=\"qrichtext\" content=\"1\" /><style type=\"text/css\">\n"
"p, li { white-space: pre-wrap; }\n"
"</style></head><body style=\" font-family:\'MS Shell Dlg 2\'; font-size:8.25pt; font-weight:400; font-style:normal;\">\n"
"<p align=\"center\" style=\" margin-top:0px; margin-bottom:0px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><span style=\" font-family:\'MS Shell Dlg 2\'; font-size:6pt; font-weight:600;\">Kernel</span></p></body></html>"))
        self.Iterations_slider.setHtml(_translate("MainWindow", "<!DOCTYPE HTML PUBLIC \"-//W3C//DTD HTML 4.0//EN\" \"http://www.w3.org/TR/REC-html40/strict.dtd\">\n"
"<html><head><meta name=\"qrichtext\" content=\"1\" /><style type=\"text/css\">\n"
"p, li { white-space: pre-wrap; }\n"
"</style></head><body style=\" font-family:\'MS Shell Dlg 2\'; font-size:8.25pt; font-weight:400; font-style:normal;\">\n"
"<p align=\"center\" style=\" margin-top:0px; margin-bottom:0px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><span style=\" font-family:\'MS Shell Dlg 2\'; font-size:6pt; font-weight:600;\">Iterations</span></p></body></html>"))
        self.ExitScreen.setText(_translate("MainWindow", "Exit"))
        self.Reset.setText(_translate("MainWindow", "Reset"))
        self.New_Master_Image.setText(_translate("MainWindow", "Import Image From DB"))
        self.customize_Image.setText(_translate("MainWindow", "Customize Image"))
        self.verify_Excel.setText(_translate("MainWindow", "Verify Excel"))
        self.Help.setText(_translate("MainWindow", "Help"))
        self.Input_Text_3.setHtml(_translate("MainWindow", "<!DOCTYPE HTML PUBLIC \"-//W3C//DTD HTML 4.0//EN\" \"http://www.w3.org/TR/REC-html40/strict.dtd\">\n"
"<html><head><meta name=\"qrichtext\" content=\"1\" /><style type=\"text/css\">\n"
"p, li { white-space: pre-wrap; }\n"
"</style></head><body style=\" font-family:\'MS Shell Dlg 2\'; font-size:8pt; font-weight:72; font-style:normal;\">\n"
"<p align=\"center\" style=\" margin-top:0px; margin-bottom:0px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><span style=\" font-size:9pt; font-weight:600;\">Cumulative Data</span></p></body></html>"))
        self.Input_Text_4.setHtml(_translate("MainWindow", "<!DOCTYPE HTML PUBLIC \"-//W3C//DTD HTML 4.0//EN\" \"http://www.w3.org/TR/REC-html40/strict.dtd\">\n"
"<html><head><meta name=\"qrichtext\" content=\"1\" /><style type=\"text/css\">\n"
"p, li { white-space: pre-wrap; }\n"
"</style></head><body style=\" font-family:\'MS Shell Dlg 2\'; font-size:8pt; font-weight:72; font-style:normal;\">\n"
"<p align=\"center\" style=\" margin-top:0px; margin-bottom:0px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><span style=\" font-weight:600;\">Present Output Data</span></p></body></html>"))
        self.menuFile.setTitle(_translate("MainWindow", "File "))
        self.menuEdit.setTitle(_translate("MainWindow", "Edit"))
        self.menuView.setTitle(_translate("MainWindow", "View"))
        self.menuSetting.setTitle(_translate("MainWindow", "Setting"))
        self.menuHelp.setTitle(_translate("MainWindow", "Help"))
        self.toolBar.setWindowTitle(_translate("MainWindow", "toolBar"))
        self.actionsave.setText(_translate("MainWindow", "New"))
        self.actionsave222.setText(_translate("MainWindow", "Open"))
        self.actionccccc.setText(_translate("MainWindow", "Save"))
        self.actionhhh.setText(_translate("MainWindow", "Print"))
        self.actionRedo.setText(_translate("MainWindow", "Redo"))
        self.actionCut.setText(_translate("MainWindow", "Cut"))
        self.actionPreference.setText(_translate("MainWindow", "Preference"))
        self.actionAction_Editor.setText(_translate("MainWindow", "Action Editor"))
        self.actionToolbar.setText(_translate("MainWindow", "Toolbar"))
        self.actionQT_GUI_Help.setText(_translate("MainWindow", "Room Address:\n"
"Sector 20, Welcome sweets junction, Atharava gym, opposite road, Matru Pitru Chaya Apt(New Apartment), Flat No 105, Airoli, Mumbai 400708\n"
""))
        self.actionRoom_Address_Sector_20_Welcome_sweets_junction_Atharava_gym_opposite_road_Matru_Pitru_Chaya_Apt_New_Apartment_Flat_No_105_Airoli_Mumbai_400708.setText(_translate("MainWindow", "Room Address:\n"
"Sector 20, Welcome sweets junction, Atharava gym, opposite road, Matru Pitru Chaya Apt(New Apartment), Flat No 105, Airoli, Mumbai 400708\n"
""))



    # ----- Functions starts -------#
    def verifyExcel(self):
        filename = QFileDialog.getOpenFileName()
        path = filename[0]

        with open(path, "r") as f:
            z=path
            # df = pd.read_excel(z)
            df = pd.read_csv(z)
            # print(df)
            print(pd.read_csv(z))
            # self.OutPutData.setText(z)
#             self.OutPutData.pd.read_csv(z)

    # To find Colour Concentration of a liquid
    def Colour_Concentration(self):
        print("Colour Concentration")

        flag_detected = 0
        Red_Counters = 0
        cap = cv2.VideoCapture(0)

        def rescale_frame(frame, percent=80):  # make the video windows a bit smaller
            width = int(frame.shape[1] * percent / 100)
            height = int(frame.shape[0] * percent / 100)
            dim = (width, height)
            return cv2.resize(frame, dim, interpolation=cv2.INTER_AREA)

        while (1 & self.aa == True):
            text = ""
            ret, frame = cap.read()
            frame = rescale_frame(frame)
            out_new = np.uint8(frame)
            out_Gray = cv2.cvtColor(out_new, cv2.COLOR_BGR2GRAY)
            ret, thresh_out = cv2.threshold(out_Gray, 127, 255, cv2.THRESH_BINARY_INV)
            kernel_ip = np.ones((2, 2), np.uint8)
            eroded_ip = cv2.erode(thresh_out, kernel_ip, iterations=1)
            dilated_ip = cv2.dilate(eroded_ip, kernel_ip, iterations=1)
            #             cv2.imshow("testing 222", dilated_ip)
            cnts = cv2.findContours(dilated_ip.copy(), cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
            cnts = imutils.grab_contours(cnts)

            if len(cnts) == 0:
                flag_empty = 1

                flag_detected = 0
                #         text = "Empty Frame"
                #         cv2.putText(frame, text, (25,25), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 0, 255),2)
                cv2.imshow("output", frame)
                cv2.waitKey(30)
                continue
            # converting  BGR to HSV Frame
            Big_faulty = max(cnts, key=cv2.contourArea)
            hsv = cv2.cvtColor(frame, cv2.COLOR_BGR2HSV)

            # the range of red color
            blu_lower = np.array([105, 142, 164], np.uint8)
            blu_upper = np.array([111, 232, 213], np.uint8)

            # finding the range of red color in the image
            red = cv2.inRange(hsv, blu_lower, blu_upper)
            kernal = np.ones((3, 3), "uint8")

            # dilation of the image ( to remove noise) create mask for red color
            red = cv2.dilate(red, kernal, iterations=1)
            res = cv2.bitwise_and(frame, frame, mask=red)

            contours, hierarchy = cv2.findContours(red, cv2.RETR_TREE, cv2.CHAIN_APPROX_SIMPLE)
            if len(contours) != 0:
                Big_Con = max(contours, key=cv2.contourArea)
                if (cv2.contourArea(Big_Con) > 5000):
                    x, y, w, h = cv2.boundingRect(Big_Con)
                    frame = cv2.rectangle(frame, (x, y), (x + w, y + h), (0, 255, 0), 2)
                    if (x > 10 & x < 420):
                        text = "concentration is good"
                # elif (cv2.contourArea(Big_faulty) > 4000):
                else:
                    x, y, w, h = cv2.boundingRect(Big_faulty)
                    frame = cv2.rectangle(frame, (x, y), (x + w, y + h), (0, 0, 255), 2)
                    if (x > 10 & x < 420):
                        if (text == "concentration is good"):
                            text = "concentration is good"
                        else:
                            text = "concentration is too high"

            cv2.putText(frame, text, (50, 50), cv2.FONT_HERSHEY_SIMPLEX, 1, (255, 255, 255))
            #cv2.imshow("output", frame)
            self.displayImage2(frame, 1)
            key = cv2.waitKey(30)
            if key == ord('q') or key == 27:
                break

    # To find the quality of cap seal on the bottle
    def Seal_Check(self):
        print("Seal Check")

        cap = cv2.VideoCapture(0)

        fourcc = cv2.VideoWriter_fourcc(*'XVID')
        out = cv2.VideoWriter('output.avi', fourcc, 30.0, (384, 512))

        def rescale_frame(frame, percent=80):  # make the video windows a bit smaller
            width = int(frame.shape[1] * percent / 100)
            height = int(frame.shape[0] * percent / 100)
            dim = (width, height)
            return cv2.resize(frame, dim, interpolation=cv2.INTER_AREA)

        while (1 & self.aa == True):
            text = ""
            ret, frame = cap.read()
            frame = rescale_frame(frame)
            out_new = np.uint8(frame)
            out_Gray = cv2.cvtColor(out_new, cv2.COLOR_BGR2GRAY)
            ret, thresh_out = cv2.threshold(out_Gray, 37, 255, cv2.THRESH_BINARY_INV)
            kernel_ip = np.ones((2, 2), np.uint8)
            eroded_ip = cv2.erode(thresh_out, kernel_ip, iterations=1)
            dilated_ip = cv2.dilate(eroded_ip, kernel_ip, iterations=1)
            #cv2.imshow("dileted", dilated_ip)
            #             cv2.imshow("testing 222", dilated_ip)
            cnts = cv2.findContours(dilated_ip.copy(), cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
            cnts = imutils.grab_contours(cnts)
            #     print(len(cnts))

            if len(cnts) == 0:
                flag_empty = 1

                flag_detected = 0
                #         text = "Empty Frame"
                #         cv2.putText(frame, text, (25,25), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 0, 255),2)
                # cv2.imshow("Decision", frame)
                self.displayImage2(frame, 1)
                cv2.waitKey(30)
                continue
            # read image and take first channel only
            # img = cv2.imread("half with cap.jpg")
            #     img = cv2.imread("stick.jpg")
            bottle_gray = cv2.cvtColor(out_new, cv2.COLOR_BGR2GRAY)
            # bottle_gray = cv2.split(bottle_3_channel)[0]
            #     cv2.imshow("Bottle Gray", bottle_gray)
            # cv2.waitKey(0)

            # blur image
            bottle_gray = cv2.GaussianBlur(bottle_gray, (7, 7), 0)
            #     cv2.imshow("Bottle Gray Smoothed 7 x 7", bottle_gray)
            # cv2.waitKey(0)
            # draw histogram
            # plt.hist(bottle_gray.ravel(), 256,[0, 256]); plt.show()

            # manual threshold
            bottle_gray = np.uint8(bottle_gray)
            bottle_threshold = cv2.threshold(bottle_gray, 20, 255, cv2.THRESH_BINARY_INV)[1]
            bottle_threshold = np.uint8(bottle_threshold)
            #     cv2.imshow("Bottle Gray Threshold 27.5", bottle_threshold)
            # cv2.waitKey(0)

            # apply opening operation
            kernel_O = cv2.getStructuringElement(cv2.MORPH_RECT, (7, 7))
            bottle_open = cv2.morphologyEx(bottle_threshold, cv2.MORPH_OPEN, kernel_O, 3)
            kernel_C = cv2.getStructuringElement(cv2.MORPH_RECT, (7, 7))
            bottle_close = cv2.morphologyEx(bottle_open, cv2.MORPH_CLOSE, kernel_C, 3)
            #cv2.imshow("Bottle Open 5 x 5", bottle_close)

            # cv2.waitKey(0)

            # find all contours
            contours = cv2.findContours(bottle_close.copy(), cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
            contours = imutils.grab_contours(contours)
            bottle_clone = out_new.copy()
            cv2.drawContours(bottle_clone, contours, -1, (0, 255, 0), 2)
            #     cv2.imshow("All Contours", bottle_clone)
            # cv2.waitKey(0)

            # sort contours by area
            areas = [cv2.contourArea(contour) for contour in contours]
            if len(areas) == 0:
                #cv2.imshow("Decision", frame)
                self.displayImage2(frame, 1)
                # print(hi)

                cv2.waitKey(30)
                continue
            (contours, areas) = zip(*sorted(zip(contours, areas), key=lambda a: a[1]))
            # print contour with largest area
            bottle_clone = out_new.copy()
            cv2.drawContours(bottle_clone, [contours[-1]], -1, (0, 255, 0), 2)
            #cv2.imshow("Largest contour", bottle_clone)
            # cv2.waitKey(0)

            # draw bounding box, calculate aspect and display decision
            bottle_clone = out_new.copy()
            (x, y, w, h) = cv2.boundingRect(contours[-1])
            # print(x,y,w,h)
            aspectRatio = w / float(h)
            # print(aspectRatio)
            print(x)
            if (60 < x < 380):
                if (aspectRatio > 3):
                    cv2.rectangle(bottle_clone, (x, y), (x + w, y + h), (0, 0, 255), 2)
                    cv2.putText(bottle_clone, "Missing Cap", (25, 25), cv2.FONT_HERSHEY_PLAIN, 1, (0, 0, 255), 2)
                elif (aspectRatio > 1.83 or aspectRatio < 1.3):
                    cv2.rectangle(bottle_clone, (x, y), (x + w, y + h), (0, 0, 255), 2)
                    cv2.putText(bottle_clone, "Open Cap", (x + 10, y + 20), cv2.FONT_HERSHEY_PLAIN, 1, (0, 0, 255), 2)
                elif (1.5 < aspectRatio < 1.63):
                    cv2.rectangle(bottle_clone, (x, y), (x + w, y + h), (0, 255, 0), 2)
                    cv2.putText(bottle_clone, "good Cap", (x + 10, y + 20), cv2.FONT_HERSHEY_PLAIN, 1, (0, 255, 0), 2)

                #cv2.imshow("Decision", bottle_clone)
                self.displayImage2(frame, 1)
                print(bottle_clone.shape)
            else:
                #cv2.imshow("Decision", frame)
                self.displayImage2(frame, 1)

            key = cv2.waitKey(30)
            if key == ord('q') or key == 27:
                break
        cap.release()
        cv2.destroyAllWindows()

    # To find the liquid level in bottle
    def Liquid_Level_Inspection(self):
        print("Liquid Level Inspection")

        cap = cv2.VideoCapture(0)

        def rescale_frame(frame, percent=80):  # make the video windows a bit smaller
            width = int(frame.shape[1] * percent / 100)
            height = int(frame.shape[0] * percent / 100)
            dim = (width, height)
            return cv2.resize(frame, dim, interpolation=cv2.INTER_AREA)

        while (1 & self.aa == True):
            text = ""
            ret, frame = cap.read()
            frame = rescale_frame(frame)
            out_new = np.uint8(frame)
            out_Gray = cv2.cvtColor(out_new, cv2.COLOR_BGR2GRAY)
            ret, thresh_out = cv2.threshold(out_Gray, 30, 255, cv2.THRESH_BINARY_INV)
            kernel_ip = np.ones((2, 2), np.uint8)
            eroded_ip = cv2.erode(thresh_out, kernel_ip, iterations=1)
            dilated_ip = cv2.dilate(eroded_ip, kernel_ip, iterations=1)
            #cv2.imshow("dileted", dilated_ip)
            #             cv2.imshow("testing 222", dilated_ip)
            cnts = cv2.findContours(dilated_ip.copy(), cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
            cnts = imutils.grab_contours(cnts)
            #     print(len(cnts))

            if len(cnts) == 0:
                flag_empty = 1

                flag_detected = 0
                #         text = "Empty Frame"
                #         cv2.putText(frame, text, (25,25), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 0, 255),2)
                #cv2.imshow("Decision", frame)
                self.displayImage2(frame, 1)
                cv2.waitKey(30)
                continue
            # read image and take first channel only
            # img = cv2.imread("half with cap.jpg")
            #     img = cv2.imread("stick.jpg")
            bottle_gray = cv2.cvtColor(out_new, cv2.COLOR_BGR2GRAY)
            # bottle_gray = cv2.split(bottle_3_channel)[0]
            #     cv2.imshow("Bottle Gray", bottle_gray)
            # cv2.waitKey(0)

            # blur image
            bottle_gray = cv2.GaussianBlur(bottle_gray, (7, 7), 0)
            #     cv2.imshow("Bottle Gray Smoothed 7 x 7", bottle_gray)
            # cv2.waitKey(0)
            # draw histogram
            # plt.hist(bottle_gray.ravel(), 256,[0, 256]); plt.show()

            # manual threshold
            bottle_gray = np.uint8(bottle_gray)
            bottle_threshold = cv2.threshold(bottle_gray, 50, 255, cv2.THRESH_BINARY_INV)[1]
            bottle_threshold = np.uint8(bottle_threshold)
            #     cv2.imshow("Bottle Gray Threshold 27.5", bottle_threshold)
            # cv2.waitKey(0)

            # apply opening operation
            kernel_O = cv2.getStructuringElement(cv2.MORPH_RECT, (7, 7))
            bottle_open = cv2.morphologyEx(bottle_threshold, cv2.MORPH_OPEN, kernel_O, 3)
            kernel_C = cv2.getStructuringElement(cv2.MORPH_RECT, (7, 7))
            bottle_close = cv2.morphologyEx(bottle_open, cv2.MORPH_CLOSE, kernel_C, 3)
            #     cv2.imshow("Bottle Open 5 x 5", bottle_close)

            # cv2.waitKey(0)

            # find all contours
            contours = cv2.findContours(bottle_close.copy(), cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
            contours = imutils.grab_contours(contours)
            bottle_clone = out_new.copy()
            cv2.drawContours(bottle_clone, contours, -1, (0, 255, 0), 2)
            #     cv2.imshow("All Contours", bottle_clone)
            # cv2.waitKey(0)

            # sort contours by area
            areas = [cv2.contourArea(contour) for contour in contours]
            if len(areas) == 0:
                #cv2.imshow("Decision", frame)
                self.displayImage2(frame, 1)
                cv2.waitKey(30)
                continue
            (contours, areas) = zip(*sorted(zip(contours, areas), key=lambda a: a[1]))
            # print contour with largest area
            bottle_clone = out_new.copy()
            cv2.drawContours(bottle_clone, [contours[-1]], -1, (0, 255, 0), 2)
            #cv2.imshow("Largest contour", bottle_clone)
            # cv2.waitKey(0)

            # draw bounding box, calculate aspect and display decision
            bottle_clone = out_new.copy()
            (x, y, w, h) = cv2.boundingRect(contours[-1])
            print(x, y, w, h)
            aspectRatio = w / float(h)
            if (80 < y < 190 and 67 < w < 110):
                cv2.putText(bottle_clone, "Too Low", (x + 10, y + 20), cv2.FONT_HERSHEY_PLAIN, 1, (0, 255, 0), 2)
                #cv2.imshow("Decision", bottle_clone)
                self.displayImage2(bottle_clone, 1)
            elif w > 125:
                if (aspectRatio < 1.5):
                    #     if ( h > 150):
                    cv2.rectangle(bottle_clone, (x, y), (x + w, y + h), (0, 255, 0), 2)
                    cv2.putText(bottle_clone, "Full", (x + 10, y + 20), cv2.FONT_HERSHEY_PLAIN, 1, (0, 255, 0), 2)

                #     elif( y+h> 155):
                #

                else:
                    cv2.rectangle(bottle_clone, (x, y), (x + w, y + h), (0, 0, 255), 2)
                    cv2.putText(bottle_clone, "Low", (x + 10, y + 20), cv2.FONT_HERSHEY_PLAIN, 1, (0, 0, 255), 2)
                #cv2.imshow("Decision", bottle_clone)
                self.displayImage2(bottle_clone, 1)
            else:
                #cv2.imshow("Decision", frame)
                self.displayImage2(frame, 1)
                cv2.waitKey(30)
                continue
            key = cv2.waitKey(30)
            if key == ord('q') or key == 27:
                break
        cap.release()
        cv2.destroyAllWindows()

    # ----- Quality check for Almond Quality ----- #
    def Almond_Quality(self):
        print("Almond Quality")

        flag_detected = 0
        Red_Counters = 0
        cap = cv2.VideoCapture(0)

        def rescale_frame(frame, percent=80):  # make the video windows a bit smaller
            width = int(frame.shape[1] * percent / 100)
            height = int(frame.shape[0] * percent / 100)
            dim = (width, height)
            return cv2.resize(frame, dim, interpolation=cv2.INTER_AREA)

        while (1 & self.aa == True):
            text = ""
            ret, frame = cap.read()
            frame = rescale_frame(frame)
            out_new = np.uint8(frame)
            out_Gray = cv2.cvtColor(out_new, cv2.COLOR_BGR2GRAY)
            ret, thresh_out = cv2.threshold(out_Gray, 127, 255, cv2.THRESH_BINARY_INV)
            kernel_ip = np.ones((2, 2), np.uint8)
            eroded_ip = cv2.erode(thresh_out, kernel_ip, iterations=1)
            dilated_ip = cv2.dilate(eroded_ip, kernel_ip, iterations=1)
            #             cv2.imshow("testing 222", dilated_ip)
            cnts = cv2.findContours(dilated_ip.copy(), cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
            cnts = imutils.grab_contours(cnts)

            if len(cnts) == 0:
                flag_empty = 1
                flag_detected = 0
                #         text = "Empty Frame"
                #         cv2.putText(frame, text, (25,25), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 0, 255),2)
                #cv2.imshow("output", frame)
                self.displayImage2(frame, 1)
                #
                cv2.waitKey(30)
                continue
            # converting  BGR to HSV Frame
            Big_faulty = max(cnts, key=cv2.contourArea)
            hsv = cv2.cvtColor(frame, cv2.COLOR_BGR2HSV)

            # the range of almond brown
            blu_lower = np.array([12, 122, 146], np.uint8)
            blu_upper = np.array([24, 192, 188], np.uint8)

            # finding the range of red color in the image
            red = cv2.inRange(hsv, blu_lower, blu_upper)
            kernal = np.ones((3, 3), "uint8")

            # dilation of the image ( to remove noise) create mask for red color
            red = cv2.dilate(red, kernal, iterations=1)
            res = cv2.bitwise_and(frame, frame, mask=red)

            contours, hierarchy = cv2.findContours(red, cv2.RETR_TREE, cv2.CHAIN_APPROX_SIMPLE)
            if len(contours) != 0:
                Big_Con = max(contours, key=cv2.contourArea)
                # print(cv2.contourArea(Big_Con))
                if (cv2.contourArea(Big_Con) > 2000):
                    x, y, w, h = cv2.boundingRect(Big_Con)
                    frame = cv2.rectangle(frame, (x, y), (x + w, y + h), (0, 255, 0), 2)
                    if (x > 10 & x < 420):
                        text = "good almond"
                # elif (cv2.contourArea(Big_faulty) > 4000):
                else:
                    x, y, w, h = cv2.boundingRect(Big_faulty)
                    frame = cv2.rectangle(frame, (x, y), (x + w, y + h), (0, 0, 255), 2)
                    if (x > 10 & x < 420):
                        if (text == "good almond"):
                            text = "good almond"
                        else:
                            text = "Bad Almond"

            cv2.putText(frame, text, (50, 50), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 0, 0))
            #cv2.imshow("output", frame)
            self.displayImage2(frame, 1)
            key = cv2.waitKey(30)
            if key == ord('q') or key == 27:
                break

    # Start here with the Main Algorithms
    def pressed(self):
            self.aa = 1
            x = self.InspectionOptions.currentText()
            self.label.setText("Option Selected: " + str(x))
            self.completed = 0
            while self.completed < 100:
                    self.completed += 0.0001
                    self.progressBar.setValue(self.completed)

            if x == "Edge Detection":
                    print("inside Edge Detection_choose")
                    self.Edge_Screen()  # edge detection
                    self.label.setText("Option Selected: " + str(x))

            if x == "Scan QR/Barcode":
                    print("inside object dimension_choose")
                    self.Scan_QR_Barcode()  # edge detection
                    self.label.setText("Option Selected: " + str(x))

            if x == "Object Dimension":
                    print("inside object dimension_choose")
                    self.Object_Dimensio()  # Object Dimensio
                    self.label.setText("Option Selected: " + str(x))

            if x == "Difference":
                    print("Selected Difference")
                    self.Difference_Object()  # Difference
                    self.label.setText("Option Selected: " + str(x))

            if x == "Colour Identification":
                    print("Selected Color Identification")
                    self.ColorWindow()
                    self.label.setText("Option Selected: " + str(x))

            if x == "Angle Detection":
                    print("Selected Angle Detection")
                    self.angle_detection()  # Angle Detection
                    self.label.setText("Option Selected: " + str(x))

            if x == "Sorting":
                    print("Selected Sorting Detection")
                    self.Sorting_detection()  # Selected Sorting Detection
                    self.label.setText("Option Selected: " + str(x))

            if x == "Multiple Color Identification":
                    print("Multiple Color Identification")
                    self.Multiple_Color_Identification()  # Multiple Color Identification
                    self.label.setText("Option Selected: " + str(x))

            if x == "Colour Concentration": # Colour Concentration
                print("Colour Concentration")
                self.Colour_Concentration()  # Colour Concentration
                self.label.setText("Option Selected: " + str(x))

            if x == "Seal Check": # Seal Check
                print("Seal Check")
                self.Seal_Check()  # Seal Check
                self.label.setText("Option Selected: " + str(x))

            if x == "Liquid Level Inspection": # Liquid Level Inspection
                print("Liquid Level Inspection")
                self.Liquid_Level_Inspection()  # Liquid Level Inspection
                self.label.setText("Option Selected: " + str(x))

            if x == "Almond Quality": # Liquid Level Inspection
                print("Almond Quality")
                self.Almond_Quality()  # Almond Quality
                self.label.setText("Option Selected: " + str(x))

            return x

    def customize_Master_Image(self):
            print("-----customize_Master_Image-----")
            vidObj = cv2.VideoCapture(0)
            while (1 & self.aa == True):
                    success, image = vidObj.read()
                    self.displayImage2(image, 1)
                    path_MV = os.getcwd()
                    cv2.imwrite(path_MV + "\\Gold\\refe.jpg", image)

    # ----- Colour Code Starts ----- #
    @QtCore.pyqtSlot(str)
    def Color_Identification(self, Select_Color):
            #
            #         In_Cam= self.Input_Camera()
            #         cap = cv2.VideoCapture(In_Cam)
            flag_detected = 0
            cap = cv2.VideoCapture(0)
            wb = xlwt.Workbook()
            sheet = wb.add_sheet('Color_Count', cell_overwrite_ok=True)
            style = xlwt.easyxf('font:bold 1')
            self.Select_Color = Select_Color

            def rescale_frame(frame, percent=80):  # make the video windows a bit smaller
                    width = int(frame.shape[1] * percent / 100)
                    height = int(frame.shape[0] * percent / 100)
                    dim = (width, height)
                    return cv2.resize(frame, dim, interpolation=cv2.INTER_AREA)

            while (1 & self.aa == True):
                    ret, frame = cap.read()
                    frame = rescale_frame(frame)
                    out_new = np.uint8(frame)
                    out_Gray = cv2.cvtColor(out_new, cv2.COLOR_BGR2GRAY)
                    ret, thresh_out = cv2.threshold(out_Gray, 127, 255, cv2.THRESH_BINARY_INV)
                    kernel_ip = np.ones((2, 2), np.uint8)
                    eroded_ip = cv2.erode(thresh_out, kernel_ip, iterations=1)
                    dilated_ip = cv2.dilate(eroded_ip, kernel_ip, iterations=1)
                    #             cv2.imshow("testing 222", dilated_ip)
                    cnts = cv2.findContours(dilated_ip.copy(), cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
                    cnts = imutils.grab_contours(cnts)
                    print(len(cnts))

                    if len(cnts) == 0:
                            flag_empty = 1

                            flag_detected = 0
                            text = "Empty Frame"
                            cv2.putText(frame, text, (25, 25), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 0, 255), 2)
                    # converting  BGR to HSV Frame
                    hsv = cv2.cvtColor(frame, cv2.COLOR_BGR2HSV)

                    if self.Select_Color == "red":
                            # the range of red color
                            red_lower = np.array([0, 87, 111], np.uint8)
                            red_upper = np.array([10, 255, 255], np.uint8)

                            # finding the range of red color in the image
                            red = cv2.inRange(hsv, red_lower, red_upper)

                            kernal = np.ones((5, 5), "uint8")

                            # dilation of the image ( to remove noise) create mask for red color
                            red = cv2.dilate(red, kernal, iterations=1)
                            res = cv2.bitwise_and(frame, frame, mask=red)

                            contours, hierarchy = cv2.findContours(red, cv2.RETR_TREE, cv2.CHAIN_APPROX_SIMPLE)
                            for pic, contour in enumerate(contours):
                                    area = cv2.contourArea(contour)
                                    if (
                                            area > 1000):  # if red color object size is grater than 1000 it will create reactangle area
                                            x, y, w, h = cv2.boundingRect(contour)
                                            frame = cv2.rectangle(frame, (x, y), (x + w, y + h), (0, 255, 0), 2)
                                            cv2.putText(frame, "PASS", (x, y), cv2.FONT_HERSHEY_SIMPLEX, 0.7,
                                                        (0, 255, 0))
                                            if flag_detected == 0:
                                                    self.Red_Counters = self.Red_Counters + 1
                                                    flag_detected = 1
                                            # self.Red_Counters = self.Red_Counters + 1

                    elif self.Select_Color == "Yellow":

                            # the range of yellow color
                            yellow_lower = np.array([22, 60, 200], np.uint8)
                            yellow_upper = np.array([60, 255, 255], np.uint8)

                            # finding the range of yellow color in the image
                            yellow = cv2.inRange(hsv, yellow_lower, yellow_upper)

                            kernal = np.ones((5, 5), "uint8")

                            # dilation of the image ( to remove noise) create mask for yellow color
                            yellow = cv2.dilate(yellow, kernal, iterations=1)
                            res = cv2.bitwise_and(frame, frame, mask=yellow)

                            contours, hierarchy = cv2.findContours(yellow, cv2.RETR_TREE, cv2.CHAIN_APPROX_SIMPLE)

                            for pic, contour in enumerate(contours):
                                    area = cv2.contourArea(contour)

                                    if area > 1000:  # if Yellow color object size is grater than 1000 it will create reactangle area
                                            x, y, w, h = cv2.boundingRect(contour)
                                            frame = cv2.rectangle(frame, (x, y), (x + w, y + h), (0, 255, 0), 2)
                                            cv2.putText(frame, "PASS", (x, y), cv2.FONT_HERSHEY_SIMPLEX, 0.7,
                                                        (0, 255, 0))
                                            if flag_detected == 0:
                                                    self.Yellow_Counters = self.Yellow_Counters + 1
                                                    flag_detected = 1
                                            # self.Yellow_Counters = self.Yellow_Counters + 1

                    elif self.Select_Color == "Blue":
                            # the range of Blue color
                            blue_lower = np.array([110, 50, 50], np.uint8)
                            blue_upper = np.array([130, 255, 255], np.uint8)

                            # finding the range of blue color in the image
                            blue = cv2.inRange(hsv, blue_lower, blue_upper)
                            kernal = np.ones((5, 5), "uint8")

                            # dilation of the image ( to remove noise) create mask for blue color
                            blue = cv2.dilate(blue, kernal, iterations=1)
                            res = cv2.bitwise_and(frame, frame, mask=blue)
                            contours, hierarchy = cv2.findContours(blue, cv2.RETR_TREE, cv2.CHAIN_APPROX_SIMPLE)

                            for pic, contour in enumerate(contours):
                                    area = cv2.contourArea(contour)
                                    if area > 1000:  # if Blue color object size is grater than 1000 it will create reactangle area
                                            x, y, w, h = cv2.boundingRect(contour)
                                            frame = cv2.rectangle(frame, (x, y), (x + w, y + h), (0, 255, 0), 2)
                                            cv2.putText(frame, "PASS", (x, y), cv2.FONT_HERSHEY_SIMPLEX, 0.7,
                                                        (0, 255, 0))
                                            if flag_detected == 0:
                                                    self.Blue_Counters = self.Blue_Counters + 1
                                                    flag_detected = 1
                    #                         self.Blue_Counters = self.Blue_Counters + 1

                    elif self.Select_Color == "Black":

                            # the range of black color
                            black_lower = np.array([0, 0, 0], np.uint8)
                            black_upper = np.array([180, 255, 30], np.uint8)

                            # finding the range of black color in the image
                            black = cv2.inRange(hsv, black_lower, black_upper)

                            kernal = np.ones((5, 5), "uint8")

                            # dilation of the image ( to remove noise) create mask for black color
                            black = cv2.dilate(black, kernal, iterations=1)
                            res = cv2.bitwise_and(frame, frame, mask=black)

                            contours, hierarchy = cv2.findContours(black, cv2.RETR_TREE, cv2.CHAIN_APPROX_SIMPLE)

                            for pic, contour in enumerate(contours):
                                    area = cv2.contourArea(contour)
                                    if (
                                            area > 1000):  # if black color object size is grater than 1000 it will create reactangle area
                                            x, y, w, h = cv2.boundingRect(contour)
                                            frame = cv2.rectangle(frame, (x, y), (x + w, y + h), (0, 255, 0), 2)
                                            cv2.putText(frame, "PASS", (x, y), cv2.FONT_HERSHEY_SIMPLEX, 0.7,
                                                        (0, 255, 0))
                                            if flag_detected == 0:
                                                    self.Black_Counters = self.Black_Counters + 1
                                                    flag_detected = 1
                                            # self.Black_Counters = self.Black_Counters + 1

                    elif self.Select_Color == "White":

                            # the range of White color
                            white_lower = np.array([0, 0, 200], np.uint8)
                            white_upper = np.array([145, 60, 255], np.uint8)

                            # finding the range of white color in the image
                            white = cv2.inRange(hsv, white_lower, white_upper)

                            kernal = np.ones((5, 5), "uint8")

                            # dilation of the image ( to remove noise) create mask for white color
                            white = cv2.dilate(white, kernal, iterations=1)
                            res = cv2.bitwise_and(frame, frame, mask=white)
                            contours, hierarchy = cv2.findContours(white, cv2.RETR_TREE, cv2.CHAIN_APPROX_SIMPLE)

                            for pic, contour in enumerate(contours):
                                    area = cv2.contourArea(contour)
                                    if (
                                            area > 1000):  # if red color object size is grater than 1000 it will create reactangle area
                                            x, y, w, h = cv2.boundingRect(contour)
                                            frame = cv2.rectangle(frame, (x, y), (x + w, y + h), (0, 255, 0), 2)
                                            cv2.putText(frame, "PASS", (x, y), cv2.FONT_HERSHEY_SIMPLEX, 0.7,
                                                        (0, 255, 0))
                                            if flag_detected == 0:
                                                    self.White_Counters = self.White_Counters + 1
                                                    flag_detected = 1
                                            # self.White_Counters = self.White_Counters + 1

                    elif self.Select_Color == "Green":

                            # the range of Green color
                            Green_lower = np.array([65, 60, 60], np.uint8)
                            Green_upper = np.array([80, 255, 255], np.uint8)

                            # finding the range of Green color in the image
                            Green = cv2.inRange(hsv, Green_lower, Green_upper)
                            kernal = np.ones((5, 5), "uint8")
                            # dilation of the image ( to remove noise) create mask for Green color
                            Green = cv2.dilate(Green, kernal, iterations=1)
                            res = cv2.bitwise_and(frame, frame, mask=Green)
                            contours, hierarchy = cv2.findContours(Green, cv2.RETR_TREE, cv2.CHAIN_APPROX_SIMPLE)

                            for pic, contour in enumerate(contours):
                                    area = cv2.contourArea(contour)
                                    if area > 1000:  # if Green color object size is grater than 1000 it will create reactangle area
                                            x, y, w, h = cv2.boundingRect(contour)
                                            frame = cv2.rectangle(frame, (x, y), (x + w, y + h), (0, 255, 0), 2)
                                            cv2.putText(frame, "PASS", (x, y), cv2.FONT_HERSHEY_SIMPLEX, 0.7,
                                                        (0, 255, 0))
                                            if flag_detected == 0:
                                                    self.Green_Counters = self.Green_Counters + 1
                                                    flag_detected = 1

                    self.displayImage2(frame, 1)

                    Color_Index = 1
                    Pass_Case_Counter = [self.Red_Counters, self.Green_Counters, self.Blue_Counters,
                                         self.Yellow_Counters,
                                         self.White_Counters, self.Black_Counters]

                    Color_List = ["Red", "Green", "Blue", "Yellow", "White", "Black"]

                    Header_Index = 0
                    Header_List = ["Sr No.", "Name Of Color", "No Of Pass Counter"]

                    # create header for all 3 coloums
                    for char in Header_List:
                            sheet.write(0, Header_Index, char, style)
                            Header_Index = Header_Index + 1

                    for i in range(len(Color_List)):
                            sheet.write(Color_Index, 0, Color_Index)
                            sheet.write(Color_Index, 1, Color_List[i])
                            sheet.write(Color_Index, 2, Pass_Case_Counter[i])
                            Color_Index = Color_Index + 1

                    wb.save('Color Identification.xls')
                    cv2.waitKey(1)

            cap.release()
            cv2.destroyAllWindows()

    def MasterImageRetrieved(self):
            print("NewMasterImage Retrieved")

            def writeTofile(data, filename):
                    # Convert binary data to proper format and write it on Hard Disk
                    with open(filename, 'wb') as file:
                            file.write(data)
                    print("Stored blob data into: ", filename, "\n")

            def readBlobData(Id):
                    try:
                            path_MV = os.getcwd()
                            sqliteConnection = sqlite3.connect(path_MV + '\\db\\SKC_MV_database.db')
                            cursor = sqliteConnection.cursor()
                            print("Connected to SQLite")

                            sql_fetch_blob_query = """SELECT * from Insert_Image where id = ?"""
                            cursor.execute(sql_fetch_blob_query, (Id,))
                            record = cursor.fetchall()
                            for row in record:
                                    print("Id = ", row[0], "Name = ", row[1])
                                    name = row[1]
                                    photo = row[2]

                                    print("Storing master image on disk \n")
                                    photoPath = path_MV + "\\master\\" + name + ".jpg"
                                    writeTofile(photo, photoPath)

                            cursor.close()

                    except sqlite3.Error as error:
                            print("Failed to read blob data from sqlite table", error)
                    finally:
                            if (sqliteConnection):
                                    sqliteConnection.close()
                                    print("sqlite connection is closed")

            textboxValue = self.lineEdit.text()
            readBlobData(textboxValue)

    # ----- Edge Screen() ----- #
    def Edge_Screen(self):

            cap = cv2.VideoCapture(0)

            def safe_div(x, y):  # so we don't crash so often
                    if y == 0: return 0
                    return x / y

            def nothing(x):  # for trackbar
                    pass

            def rescale_frame(frame, percent=80):  # make the video windows a bit smaller
                    width = int(frame.shape[1] * percent / 100)
                    height = int(frame.shape[0] * percent / 100)
                    dim = (width, height)
                    return cv2.resize(frame, dim, interpolation=cv2.INTER_AREA)

            if not cap.isOpened():
                    print("can't open camera")
                    exit()

            showLive = True
            while ((cap.isOpened() & self.aa == True)):
                    text = ""

                    ret, frame = cap.read()
                    frame = rescale_frame(frame)
                    fshape = frame.shape
                    fheight = fshape[0]
                    fwidth = fshape[1]
                    #     frame_resize = rescale_frame(frame)
                    if not ret:
                            print("cannot capture the frame")
                            exit()
                    out_Gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
                    ret, thresh_out = cv2.threshold(out_Gray, 127, 255, cv2.THRESH_BINARY_INV)
                    cnts = cv2.findContours(thresh_out, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
                    cnts = imutils.grab_contours(cnts)

                    if len(cnts) == 0:
                            text = "frame is empty"
                            cv2.putText(frame, text, (25, 25), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 0, 255), 2)
                            self.displayImage2(frame, 1)
                            #                 cv2.imshow(windowName, frame)
                            cv2.waitKey(10)
                            continue
                    # changing slider value for Thresold,Kernal,Iteration
                    Thresh = self.Thrash_slider.value()
                    kern = self.Kernel_slider.value()
                    iter1 = self.horizontalSlider_3.value()

                    # show value of threshold,kernal and Iteration to label
                    self.Threshold_Value.setNum(Thresh)
                    self.Kernel_value.setNum(kern)
                    self.Iteration_Value.setNum(iter1)
                    ret, thresh1 = cv2.threshold(frame, Thresh, 255, cv2.THRESH_BINARY_INV)
                    kernel = np.ones((kern, kern), np.uint8)  # square image kernel used for erosion
                    dilation = cv2.dilate(thresh1, kernel, iterations=iter1)
                    erosion = cv2.erode(dilation, kernel, iterations=iter1)  # refines all edges in the binary image

                    opening = cv2.morphologyEx(erosion, cv2.MORPH_OPEN, kernel)
                    closing = cv2.morphologyEx(opening, cv2.MORPH_CLOSE, kernel)
                    closing = cv2.cvtColor(closing, cv2.COLOR_BGR2GRAY)
                    contours, hierarchy = cv2.findContours(closing, cv2.RETR_TREE,
                                                           cv2.CHAIN_APPROX_NONE)  # find contours with simple approximation cv2.RETR_TREE,cv2.CHAIN_APPROX_SIMPLE
                    closing = cv2.cvtColor(closing, cv2.COLOR_GRAY2RGB)
                    cv2.drawContours(closing, contours, -1, (128, 255, 0), 1)

                    # focus on only the largest outline by area
                    areas = []  # list to hold all areas

                    for contour in contours:
                            ar = cv2.contourArea(contour)
                            areas.append(ar)

                    max_area = max(areas)
                    max_area_index = areas.index(max_area)  # index of the list element with largest area
                    cnt = contours[max_area_index]  # largest area contour is usually the viewing window itself, why?
                    cv2.drawContours(closing, [cnt], 0, (0, 255, 255), 1)

                    def midpoint(ptA, ptB):
                            return ((ptA[0] + ptB[0]) * 0.5, (ptA[1] + ptB[1]) * 0.5)

                    # compute the rotated bounding box of the contour
                    orig = frame.copy()
                    box = cv2.minAreaRect(cnt)
                    box = cv2.cv.BoxPoints(box) if imutils.is_cv2() else cv2.boxPoints(box)
                    box = np.array(box, dtype="int")

                    # order the points in the contour such that they appear
                    # in top-left, top-right, bottom-right, and bottom-left
                    # order, then draw the outline of the rotated bounding
                    # box
                    box = perspective.order_points(box)
                    cv2.drawContours(orig, [box.astype("int")], -1, (0, 255, 0), 3)

                    # loop over the original points and draw them
                    for (x, y) in box:
                            cv2.circle(orig, (int(x), int(y)), 5, (0, 0, 255), -1)

                    # unpack the ordered bounding box, then compute the midpoint
                    # between the top-left and top-right coordinates, followed by
                    # the midpoint between bottom-left and bottom-right coordinates
                    (tl, tr, br, bl) = box
                    (tltrX, tltrY) = midpoint(tl, tr)
                    (blbrX, blbrY) = midpoint(bl, br)

                    # compute the midpoint between the top-left and top-right points,
                    # followed by the midpoint between the top-righ and bottom-right
                    (tlblX, tlblY) = midpoint(tl, bl)
                    (trbrX, trbrY) = midpoint(tr, br)

                    # draw the midpoints on the image
                    cv2.circle(orig, (int(tltrX), int(tltrY)), 5, (255, 0, 0), -1)
                    cv2.circle(orig, (int(blbrX), int(blbrY)), 5, (255, 0, 0), -1)
                    cv2.circle(orig, (int(tlblX), int(tlblY)), 5, (255, 0, 0), -1)
                    cv2.circle(orig, (int(trbrX), int(trbrY)), 5, (255, 0, 0), -1)

                    # draw lines between the midpoints
                    cv2.line(orig, (int(tltrX), int(tltrY)), (int(blbrX), int(blbrY)), (255, 0, 255), 1)
                    cv2.line(orig, (int(tlblX), int(tlblY)), (int(trbrX), int(trbrY)), (255, 0, 255), 1)
                    cv2.drawContours(orig, [cnt], 0, (0, 0, 255), 1)
                    self.displayImage2(orig, 1)
                    cv2.waitKey(10)

            cap.release()
            cv2.destroyAllWindows()

    # ----- Object Difference ----- #
    def Multiple_Color_Identification(self):
            print("Multiple Color Identification")
            #         In_Cam= self.Input_Camera()
            #         cap = cv2.VideoCapture(In_Cam)
            cap = cv2.VideoCapture(0)
            wb = xlwt.Workbook()
            sheet = wb.add_sheet('Color_Count', cell_overwrite_ok=True)
            style = xlwt.easyxf('font:bold 1')

            def rescale_frame(frame, percent=80):  # make the video windows a bit smaller
                    width = int(frame.shape[1] * percent / 100)
                    height = int(frame.shape[0] * percent / 100)
                    dim = (width, height)
                    return cv2.resize(frame, dim, interpolation=cv2.INTER_AREA)

            while (1 & self.aa == True):
                    ret, frame = cap.read()
                    frame = rescale_frame(frame)
                    out_new = np.uint8(frame)
                    out_Gray = cv2.cvtColor(out_new, cv2.COLOR_BGR2GRAY)
                    ret, thresh_out = cv2.threshold(out_Gray, 127, 255, cv2.THRESH_BINARY_INV)
                    kernel_ip = np.ones((2, 2), np.uint8)
                    eroded_ip = cv2.erode(thresh_out, kernel_ip, iterations=1)
                    dilated_ip = cv2.dilate(eroded_ip, kernel_ip, iterations=1)
                    #             cv2.imshow("testing 222", dilated_ip)
                    cnts = cv2.findContours(dilated_ip.copy(), cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
                    cnts = imutils.grab_contours(cnts)
                    print(len(cnts))

                    if len(cnts) == 0:
                            flag_empty = 1
                            flag_detected = 0
                            text = "Empty Frame"
                            cv2.putText(frame, text, (25, 25), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 255, 0))
                    #             self.displayImage(frame, 1)
                    # converting  BGR to HSV Frame
                    hsv = cv2.cvtColor(frame, cv2.COLOR_BGR2HSV)

                    # the range of red color
                    red_lower = np.array([0, 87, 111], np.uint8)
                    red_upper = np.array([10, 255, 255], np.uint8)

                    # the range of yellow color
                    yellow_lower = np.array([22, 60, 200], np.uint8)
                    yellow_upper = np.array([60, 255, 255], np.uint8)

                    # the range of Green color
                    Green_lower = np.array([36, 25, 25], np.uint8)
                    Green_upper = np.array([70, 255, 255], np.uint8)

                    # the range of Blue color
                    blue_lower = np.array([110, 100, 100], np.uint8)
                    blue_upper = np.array([130, 255, 255], np.uint8)

                    # the range of black color
                    black_lower = np.array([0, 0, 0], np.uint8)
                    black_upper = np.array([180, 255, 30], np.uint8)

                    # the range of White color
                    white_lower = np.array([0, 0, 0], np.uint8)
                    white_upper = np.array([0, 0, 255], np.uint8)

                    # finding the range of red color in the image
                    red = cv2.inRange(hsv, red_lower, red_upper)

                    # finding the range of yellow color in the image
                    yellow = cv2.inRange(hsv, yellow_lower, yellow_upper)

                    # finding the range of blue color in the image
                    blue = cv2.inRange(hsv, blue_lower, blue_upper)

                    # finding the range of black color in the image
                    black = cv2.inRange(hsv, black_lower, black_upper)

                    # finding the range of white color in the image
                    white = cv2.inRange(hsv, white_lower, white_upper)

                    # finding the range of Green color in the image
                    Green = cv2.inRange(hsv, Green_lower, Green_upper)

                    kernal = np.ones((5, 5), "uint8")

                    # dilation of the image ( to remove noise) create mask for red color-------------------------------------
                    red = cv2.dilate(red, kernal, iterations=1)
                    res = cv2.bitwise_and(frame, frame, mask=red)

                    contours, hierarchy = cv2.findContours(red, cv2.RETR_TREE, cv2.CHAIN_APPROX_SIMPLE)
                    for pic, contour in enumerate(contours):
                            area = cv2.contourArea(contour)
                            if area > 5000:  # if red color object size is grater than 1000 it will create reactangle area
                                    x, y, w, h = cv2.boundingRect(contour)
                                    frame = cv2.rectangle(frame, (x, y), (x + w, y + h), (0, 0, 255), 2)
                                    cv2.putText(frame, "Red", (x, y), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 0, 255))
                                    if flag_detected == 0:
                                            self.Red_Counters = self.Red_Counters + 1
                                            flag_detected = 1
                                    # self.Red_Counters = self.Red_Counters + 1

                    # dilation of the image ( to remove noise) create mask for yellow color------------------------------------
                    yellow = cv2.dilate(yellow, kernal, iterations=1)
                    res = cv2.bitwise_and(frame, frame, mask=yellow)

                    contours, hierarchy = cv2.findContours(yellow, cv2.RETR_TREE, cv2.CHAIN_APPROX_SIMPLE)
                    for pic, contour in enumerate(contours):
                            area = cv2.contourArea(contour)
                            if area > 5000:  # if Yellow color object size is grater than 1000 it will create reactangle area
                                    x, y, w, h = cv2.boundingRect(contour)
                                    frame = cv2.rectangle(frame, (x, y), (x + w, y + h), (0, 255, 255), 2)
                                    cv2.putText(frame, "Yellow", (x, y), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 255, 255))
                                    if flag_detected == 0:
                                            self.Yellow_Counters = self.Yellow_Counters + 1
                                            flag_detected = 1
                            # self.Yellow_Counters = self.Yellow_Counters + 1

                    # dilation of the image ( to remove noise) create mask for blue color-----------------------------------
                    blue = cv2.dilate(blue, kernal, iterations=1)
                    res = cv2.bitwise_and(frame, frame, mask=blue)
                    contours, hierarchy = cv2.findContours(blue, cv2.RETR_TREE, cv2.CHAIN_APPROX_SIMPLE)

                    for pic, contour in enumerate(contours):
                            area = cv2.contourArea(contour)
                            if area > 5000:  # if Blue color object size is grater than 1000 it will create reactangle area
                                    x, y, w, h = cv2.boundingRect(contour)
                                    frame = cv2.rectangle(frame, (x, y), (x + w, y + h), (255, 0, 0), 2)
                                    cv2.putText(frame, "Blue", (x, y), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (255, 0, 0))
                                    if flag_detected == 0:
                                            self.Blue_Counters = self.Blue_Counters + 1
                                            flag_detected = 1
                                    # self.Blue_Counters = self.Blue_Counters + 1

                    # dilation of the image ( to remove noise) create mask for black color-------------------------------------
                    black = cv2.dilate(black, kernal, iterations=1)
                    res = cv2.bitwise_and(frame, frame, mask=black)
                    contours, hierarchy = cv2.findContours(black, cv2.RETR_TREE, cv2.CHAIN_APPROX_SIMPLE)

                    for pic, contour in enumerate(contours):
                            area = cv2.contourArea(contour)
                            if area > 5000:  # if black color object size is grater than 1000 it will create reactangle area
                                    x, y, w, h = cv2.boundingRect(contour)
                                    frame = cv2.rectangle(frame, (x, y), (x + w, y + h), (0, 0, 0), 2)
                                    cv2.putText(frame, "Black", (x, y), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 0, 0))
                                    if flag_detected == 0:
                                            self.Black_Counters = self.Black_Counters + 1
                                            flag_detected = 1
                                    # self.Black_Counters = self.Black_Counters + 1

                    # dilation of the image ( to remove noise) create mask for white color-----------------------------------------
                    white = cv2.dilate(white, kernal, iterations=1)
                    res = cv2.bitwise_and(frame, frame, mask=white)
                    contours, hierarchy = cv2.findContours(white, cv2.RETR_TREE, cv2.CHAIN_APPROX_SIMPLE)

                    for pic, contour in enumerate(contours):
                            area = cv2.contourArea(contour)
                            if area > 5000:  # if white color object size is grater than 1000 it will create reactangle area
                                    x, y, w, h = cv2.boundingRect(contour)
                                    frame = cv2.rectangle(frame, (x, y), (x + w, y + h), (255, 255, 255), 2)
                                    cv2.putText(frame, "White", (x, y), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (255, 255, 255))
                                    if flag_detected == 0:
                                            self.White_Counters = self.White_Counters + 1
                                            flag_detected = 1
                                    # self.White_Counters = self.White_Counters + 1

                    # dilation of the image ( to remove noise) create mask for Green color---------------------------------------------
                    Green = cv2.dilate(Green, kernal, iterations=1)
                    res = cv2.bitwise_and(frame, frame, mask=Green)
                    contours, hierarchy = cv2.findContours(Green, cv2.RETR_TREE, cv2.CHAIN_APPROX_SIMPLE)

                    for pic, contour in enumerate(contours):
                            area = cv2.contourArea(contour)
                            if area > 5000:  # if Green color object size is grater than 1000 it will create reactangle area
                                    x, y, w, h = cv2.boundingRect(contour)
                                    frame = cv2.rectangle(frame, (x, y), (x + w, y + h), (0, 255, 0), 2)
                                    cv2.putText(frame, "Green", (x, y), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 255, 0))
                                    if flag_detected == 0:
                                            self.Green_Counters = self.Green_Counters + 1
                                            flag_detected = 1
                                    # self.Green_Counters = self.Green_Counters + 1

                    self.displayImage2(frame, 1)
                    Color_Index = 1
                    Pass_Case_Counter = [self.Red_Counters, self.Green_Counters, self.Blue_Counters,
                                         self.Yellow_Counters,
                                         self.White_Counters, self.Black_Counters]

                    Color_List = ["Red", "Green", "Blue", "Yellow", "White", "Black"]
                    Header_Index = 0
                    Header_List = ["Sr No.", "Name Of Color", "Counter"]

                    # create header for all 3 coloums
                    for char in Header_List:
                            sheet.write(0, Header_Index, char, style)
                            Header_Index = Header_Index + 1

                    for i in range(len(Color_List)):
                            sheet.write(Color_Index, 0, Color_Index)
                            sheet.write(Color_Index, 1, Color_List[i])
                            sheet.write(Color_Index, 2, Pass_Case_Counter[i])
                            Color_Index = Color_Index + 1

                    wb.save('Color Identification_Multiple.xls')
                    cv2.waitKey(1)

            cap.release()
            cv2.destroyAllWindows()

    # ----- Object Difference ----- #
    def Difference_Object(self):
            print("Object Difference")

            def pad_images_to_same_size(images):
                    width_max = 0
                    height_max = 0
                    for img in images:
                            h, w = img.shape[:2]
                            width_max = max(width_max, w)
                            height_max = max(height_max, h)

                    images_padded = []
                    for img in images:
                            h, w = img.shape[:2]
                            diff_vert = height_max - h
                            pad_top = diff_vert // 2
                            pad_bottom = diff_vert - pad_top
                            diff_hori = width_max - w
                            pad_left = diff_hori // 2
                            pad_right = diff_hori - pad_left
                            img_padded = cv2.copyMakeBorder(img, pad_top, pad_bottom, pad_left, pad_right,
                                                            cv2.BORDER_CONSTANT, value=0)
                            assert img_padded.shape[:2] == (height_max, width_max)
                            images_padded.append(img_padded)

                    return images_padded

            def masterimg(image):
                    # load the image, convert it to grayscale, and blur it slightly
                    image = cv2.imread(image)
                    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
                    gray = cv2.GaussianBlur(gray, (5, 5), 0)
                    # threshold the image, then perform a series of erosions + dilations to remove any small regions of noise
                    thresh = cv2.threshold(gray, 160, 255, cv2.THRESH_BINARY_INV)[1]
                    thresh = cv2.erode(thresh, None, iterations=2)
                    thresh = cv2.dilate(thresh, None, iterations=2)
                    # find contours in thresholded image, then grab the largest one
                    cnts = cv2.findContours(thresh.copy(), cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
                    cnts = imutils.grab_contours(cnts)
                    c = max(cnts, key=cv2.contourArea)
                    #     THIS IS ADDED PART
                    rect = cv2.minAreaRect(c)
                    box = cv2.boxPoints(rect)
                    box = np.int0(box)
                    image = cv2.drawContours(image, [box], 0, (0, 0, 255), 2)
                    i = 0

                    for p in box:
                            pt = (p[0], p[1])
                            i = i + 1
                            cv2.circle(image, pt, 6, (200, 0, 0), -6)
                            if i == 1:
                                    print('Ext bottom_m : ', pt)
                                    Ext_bottom = pt
                            if i == 2:
                                    print('Ext left_m : ', pt)
                                    Ext_left = pt
                            if i == 3:
                                    print('Ext top_m : ', pt)
                                    Ext_top = pt
                            if i == 4:
                                    print('Ext right_m : ', pt)
                                    Ext_right = pt

                    cv2.drawContours(image, [c], -1, (0, 255, 255), 2)
                    cv2.circle(image, Ext_left, 6, (0, 0, 255), -1)
                    cv2.circle(image, Ext_right, 6, (0, 255, 0), -1)
                    cv2.circle(image, Ext_top, 6, (255, 0, 0), -1)
                    cv2.circle(image, Ext_bottom, 6, (255, 255, 0), -1)

                    #   defining points
                    a_m = Ext_left
                    c_m = Ext_top

                    #   Assigning cordinates
                    Ext_left_x = a_m[0]
                    Ext_top_y = c_m[1]
                    print('ExtLeft_x =', Ext_left_x)

                    global Le_range
                    Le_range = Ext_left_x  # giving reference point from Master Image
                    print('L_range = ', Le_range)

                    global Ri_range
                    Ri_range = Ext_left_x + 10  # giving reference point from Master Image
                    print('R_range = ', Ri_range)

                    global To_range
                    To_range = Ext_top_y + 50
                    print('T_range = ', To_range)

                    global Bo_range
                    Bo_range = Ext_top_y - 50
                    print('B_range = ', Bo_range)

            def create_connection(db_file):
                    """ create a database connection to the SQLite database
                        specified by db_file
                    :param db_file: database file
                    :return: Connection object or None
                    """
                    conn = None
                    try:
                            conn = sqlite3.connect(db_file)
                    except Error as e:
                            print(e)

                    return conn

            def create_task(conn, task):
                    """
                    Create a new task
                    :param conn:
                    :param task:
                    :return:
                    """
                    sql = ''' INSERT INTO Image_Difference(Total,Passed,Failed)
                      VALUES(?,?,?) '''
                    cur = conn.cursor()
                    cur.execute(sql, task)
                    return cur.lastrowid

            def writer():
                    stamp_count = 0
                    csv.write("{},{},{}\n".format("timestamp", "case ", "status"))
                    csv.flush()
                    path_MV = os.getcwd()
                    # create a database connection
                    database = path_MV + '\\db\\SKC_MV_database.db'
                    conn = create_connection(database)
                    for key in found.keys():
                            csv.write("{},{},{}\n".format(stamp[stamp_count], key, found[key]))
                            stamp_count += 1
                            csv.flush()
                    csv.write("\n")
                    csv.write("{},{}\n".format("total cases", pass_counter + fail_counter))
                    csv.write("{},{}\n".format("total passed cases", pass_counter))
                    csv.write("{},{}\n".format("total failed cases", fail_counter))
                    task_1 = (pass_counter + fail_counter, pass_counter, fail_counter)
                    with conn:
                            create_task(conn, task_1)
                    csv.flush()
                    csv.close()

            def pad_images_to_same_size(images):

                    width_max = 0
                    height_max = 0
                    for img in images:
                            h, w = img.shape[:2]
                            width_max = max(width_max, w)
                            height_max = max(height_max, h)

                    images_padded = []
                    for img in images:
                            h, w = img.shape[:2]
                            diff_vert = height_max - h
                            pad_top = diff_vert // 2
                            pad_bottom = diff_vert - pad_top
                            diff_hori = width_max - w
                            pad_left = diff_hori // 2
                            pad_right = diff_hori - pad_left
                            img_padded = cv2.copyMakeBorder(img, pad_top, pad_bottom, pad_left, pad_right,
                                                            cv2.BORDER_CONSTANT, value=0)
                            assert img_padded.shape[:2] == (height_max, width_max)
                            images_padded.append(img_padded)

                    return images_padded

            def Four_point_transform_main(img):

                    frame = cv2.resize(img, None, fx=1.0, fy=1.0, interpolation=cv2.INTER_AREA)
                    gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
                    gray = cv2.GaussianBlur(gray, (5, 5), 0)

                    # threshold the image, then perform a series of erosions +
                    # dilations to remove any small regions of noise
                    thresh = cv2.threshold(gray, 170, 255, cv2.THRESH_BINARY_INV)[1]
                    thresh = cv2.erode(thresh, None, iterations=2)
                    thresh = cv2.dilate(thresh, None, iterations=2)
                    xyz = extreme_points_K(thresh, frame)

                    # construct the argument parse and parse the arguments
                    ap = argparse.ArgumentParser()
                    ap.add_argument("-i", "--frame", help="path to the image file")
                    ap.add_argument("-c", "--coords",
                                    help="comma seperated list of source points")
                    args = vars(ap.parse_args())
                    pts = xyz
                    # apply the four point tranform to obtain a "birds eye view" of the image
                    warped = four_point_transform(frame, pts)
                    cv2.waitKey(1)
                    return warped

            def order_points(pts):
                    # initialzie a list of coordinates that will be ordered
                    # such that the first entry in the list is the top-left,
                    # the second entry is the top-right, the third is the
                    # bottom-right, and the fourth is the bottom-left
                    rect = np.zeros((4, 2), dtype="float32")

                    # the top-left point will have the smallest sum, whereas
                    # the bottom-right point will have the largest sum
                    s = pts.sum(axis=1)
                    rect[0] = pts[np.argmin(s)]
                    rect[2] = pts[np.argmax(s)]

                    # now, compute the difference between the points, the
                    # top-right point will have the smallest difference,
                    # whereas the bottom-left will have the largest difference
                    diff = np.diff(pts, axis=1)
                    rect[1] = pts[np.argmin(diff)]
                    rect[3] = pts[np.argmax(diff)]

                    # return the ordered coordinates
                    return rect

            def four_point_transform(image, pts):
                    # obtain a consistent order of the points and unpack them
                    # individually
                    rect = order_points(pts)
                    (tl, tr, br, bl) = rect

                    # compute the width of the new image, which will be the
                    # maximum distance between bottom-right and bottom-left
                    # x-coordiates or the top-right and top-left x-coordinates
                    widthA = np.sqrt(((br[0] - bl[0]) ** 2) + ((br[1] - bl[1]) ** 2))
                    widthB = np.sqrt(((tr[0] - tl[0]) ** 2) + ((tr[1] - tl[1]) ** 2))
                    maxWidth = max(int(widthA), int(widthB))

                    # compute the height of the new image, which will be the
                    # maximum distance between the top-right and bottom-right
                    # y-coordinates or the top-left and bottom-left y-coordinates
                    heightA = np.sqrt(((tr[0] - br[0]) ** 2) + ((tr[1] - br[1]) ** 2))
                    heightB = np.sqrt(((tl[0] - bl[0]) ** 2) + ((tl[1] - bl[1]) ** 2))
                    maxHeight = max(int(heightA), int(heightB))

                    # now that we have the dimensions of the new image, construct
                    # the set of destination points to obtain a "birds eye view",
                    # (i.e. top-down view) of the image, again specifying points
                    # in the top-left, top-right, bottom-right, and bottom-left
                    # order
                    dst = np.array([
                            [0, 0],
                            [maxWidth - 1, 0],
                            [maxWidth - 1, maxHeight - 1],
                            [0, maxHeight - 1]], dtype="float32")

                    # compute the perspective transform matrix andiffprocessord then apply it
                    M = cv2.getPerspectiveTransform(rect, dst)
                    warped = cv2.warpPerspective(image, M, (maxWidth, maxHeight))

                    # return the warped image
                    return warped

            def extreme_points_K(thresh, image):
                    # find contours in thresholded image, then grab the largest one
                    cnts = cv2.findContours(thresh.copy(), cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
                    cnts = imutils.grab_contours(cnts)
                    c = max(cnts, key=cv2.contourArea)
                    rect = cv2.minAreaRect(c)
                    box = cv2.boxPoints(rect)
                    box = np.int0(box)
                    return box

            # allign image function defination
            def alignImages(im1, im2):

                    im1Gray = cv2.cvtColor(im1, cv2.COLOR_BGR2GRAY)
                    im2Gray = cv2.cvtColor(im2, cv2.COLOR_BGR2GRAY)

                    # Detect ORB features and compute descriptors.
                    orb = cv2.ORB_create(MAX_FEATURES)
                    keypoints1, descriptors1 = orb.detectAndCompute(im1Gray, None)
                    keypoints2, descriptors2 = orb.detectAndCompute(im2Gray, None)

                    # Match features.
                    matcher = cv2.DescriptorMatcher_create(cv2.DESCRIPTOR_MATCHER_BRUTEFORCE_HAMMING)
                    matches = matcher.match(descriptors1, descriptors2, None)

                    # Sort matches by score
                    matches.sort(key=lambda x: x.distance, reverse=False)

                    # Remove not so good matches
                    numGoodMatches = int(len(matches) * GOOD_MATCH_PERCENT)
                    matches = matches[:numGoodMatches]

                    # Draw top matches
                    imMatches = cv2.drawMatches(im1, keypoints1, im2, keypoints2, matches, None)
                    # cv2.imwrite("matches.jpg", imMatches)

                    # Extract location of good matches
                    points1 = np.zeros((len(matches), 2), dtype=np.float32)
                    points2 = np.zeros((len(matches), 2), dtype=np.float32)

                    for i, match in enumerate(matches):
                            points1[i, :] = keypoints1[match.queryIdx].pt
                            points2[i, :] = keypoints2[match.trainIdx].pt

                    # Find homography
                    h, mask = cv2.findHomography(points1, points2, cv2.RANSAC)

                    # Use homography
                    height, width, channels = im2.shape
                    im1Reg = cv2.warpPerspective(im1, h, (width, height))
                    flag_imagecaptred = 0
                    return im1Reg

            def filterImg_avi(image):
                    # print("Enter_Filter_Image")
                    image = cv2.resize(image, None, fx=1.0, fy=1.0, interpolation=cv2.INTER_AREA)
                    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
                    pic_arr = np.asarray(image)
                    pic_arr01 = np.asarray(gray)
                    img = pic_arr01
                    gray = cv2.GaussianBlur(img, (5, 5), 0)
                    # threshold the image, then perform a series of erosions + dilations to remove any small regions of noise
                    thresh = cv2.threshold(gray, 160, 255, cv2.THRESH_BINARY_INV)[1]
                    thresh = cv2.erode(thresh, None, iterations=2)
                    thresh = cv2.dilate(thresh, None, iterations=2)
                    # find contours in thresholded image, then grab the largest one
                    cnts = cv2.findContours(thresh.copy(), cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
                    cnts = imutils.grab_contours(cnts)
                    if len(cnts) == 0:
                            active_avi = 1
                            flag_imagecaptred = 0
                            return image, flag_imagecaptred, active_avi

                    c = max(cnts, key=cv2.contourArea)
                    # THIS IS ADDED PART
                    rect = cv2.minAreaRect(c)
                    box = cv2.boxPoints(rect)
                    box = np.int0(box)
                    img = cv2.drawContours(img, [box], 0, (0, 0, 255), 2)
                    i = 0

                    for p in box:
                            pt = (p[0], p[1])
                            i = i + 1
                            cv2.circle(img, pt, 6, (200, 0, 0), -6)
                            if i == 1:
                                    Ext_bottom = pt
                            if i == 2:
                                    Ext_left = pt
                            if i == 3:
                                    Ext_top = pt
                            if i == 4:
                                    Ext_right = pt

                    cv2.drawContours(img, [c], -1, (0, 255, 255), 2)
                    cv2.circle(img, Ext_left, 6, (0, 0, 255), -1)
                    cv2.circle(img, Ext_right, 6, (0, 255, 0), -1)
                    cv2.circle(img, Ext_top, 6, (255, 0, 0), -1)
                    cv2.circle(img, Ext_bottom, 6, (255, 255, 0), -1)
                    a_p = Ext_left
                    c_p = Ext_top
                    extLeft_x_p = a_p[0]
                    print("extLeft_x_p: ", extLeft_x_p)
                    L_range = Le_range
                    R_range = Ri_range
                    T_range = To_range
                    B_range = Bo_range
                    if (L_range < extLeft_x_p and extLeft_x_p < R_range):
                            # and B_range < exttop_Y_p and exttop_Y_p < T_range
                            image = pic_arr
                            active_avi = 0
                            flag_imagecaptred = 1
                    else:
                            # print("Not equal")
                            flag_imagecaptred = 0
                            active_avi = 1
                    # print('-----Further processing images are filtered from bulk of images-----')

                    cv2.waitKey(10)
                    return image, flag_imagecaptred, active_avi

            def diffprocessor(img1, img2):
                    text = ""
                    im1 = img1
                    fault_save = im1.copy()
                    # cv2.imshow("image 1", im1)
                    Gray_im1 = cv2.cvtColor(im1, cv2.COLOR_BGR2GRAY)
                    ret, thresh_im1 = cv2.threshold(Gray_im1, 117, 255, cv2.THRESH_BINARY_INV)
                    #             cv2.imshow("thresh_im1",thresh_im1)
                    im2 = img2
                    # cv2.imshow("image 2", im2)
                    Gray_im2 = cv2.cvtColor(im2, cv2.COLOR_BGR2GRAY)
                    ret, thresh_im2 = cv2.threshold(Gray_im2, 117, 255, cv2.THRESH_BINARY_INV)
                    #             cv2.imshow("thresh_im2",thresh_im2)
                    # diff = cv2.absdiff(thresh_im1, thresh_im2)
                    diff = thresh_im2 - thresh_im1
                    #             cv2.imshow("threshold_DIFF", diff)
                    kernel = np.ones((3, 3), np.uint8)
                    erosion = cv2.erode(diff, kernel, iterations=1)
                    dilation = cv2.dilate(erosion, kernel, iterations=1)
                    out_new = np.uint8(dilation)

                    # canny = cv2.Canny(thresh_out, 120, 255)
                    canny = cv2.Canny(out_new, 120, 255)

                    ## find the non-zero min-max coords of canny
                    pts = np.argwhere(canny > 5)
                    print(len(pts))
                    if len(pts) != 0:
                            y1, x1 = pts.min(axis=0)
                            y2, x2 = pts.max(axis=0)
                            tagged = cv2.rectangle(fault_save, (x1, y1), (x2, y2), (0, 0, 200), 2, cv2.LINE_AA)
                            if x2 - x1 > 5:
                                    text = "Falty PCB"

                                    # cv2.putText(tagged, "FAIL CASE!!", (x1, y1 - 10), cv2.FONT_HERSHEY_SIMPLEX, 2, (0, 0, 255), 2)
                                    path_MV = os.getcwd()
                                    cv2.imwrite(path_MV + "//frames//frames%d.jpg" % count, tagged)
                                    cv2.putText(tagged, text, (100, 100), cv2.FONT_HERSHEY_SIMPLEX, 2, (0, 0, 255), 2)

                    else:
                            # tagged = output_ImageB.copy()
                            text = "Good PCB"

                    active_diff = 0
                    flag_imagecaptred = 0
                    return text, flag_imagecaptred, active_diff

            def rescale_frame(frame, percent=80):  # make the video windows a bit smaller
                    width = int(frame.shape[1] * percent / 100)
                    height = int(frame.shape[0] * percent / 100)
                    dim = (width, height)
                    return cv2.resize(frame, dim, interpolation=cv2.INTER_AREA)

            filename = datetime.datetime.now()
            a = int(filename.strftime('%d'))
            csv = open(filename.strftime("%d %B %Y-%Hh%Mm%Ss") + ".csv", "w")
            global pic_arr01
            path_MV = os.getcwd()
            refe = cv2.imread(path_MV + "\\master\\refe.jpg")
            ref_trans = refe.copy()
            cropped_refe = Four_point_transform_main(ref_trans)
            path_MV = os.getcwd()
            masterimg(path_MV + "\\master\\refe.jpg")
            ##for feature exxtraction
            MAX_FEATURES = 1000
            GOOD_MATCH_PERCENT = 0.5
            flag_empty = 0
            flag_imagecaptred = 0
            active_avi = 0
            active_diff = 0
            count = 0

            ####for writing in csv file
            csv_count = 0
            flag_write = 0
            found = {}
            stamp = []
            breakloop = 0
            pass_counter = 0
            fail_counter = 0
            text = ""
            ####for writing in csv file###end
            vidObj = cv2.VideoCapture(0)
            while (1 & self.aa == True):

                    latest = datetime.datetime.now()
                    b = int(latest.strftime('%d'))

                    # check if date is changed, if yes write the data and close the
                    # old csv and then create the new one
                    if (a != b):
                            writer()
                            csv = open(latest.strftime("%d %B %Y-%Hh%Mm%Ss") + ".csv", "w")
                            a = b

                    text = " "
                    success, image = vidObj.read()
                    image = rescale_frame(image)
                    out_img = image.copy()
                    if success == 0:
                            print("No Video to process")
                            quit()
                    if (success == True & self.aa == True):
                            count += 1

                            out_new = np.uint8(image)
                            out_Gray = cv2.cvtColor(out_new, cv2.COLOR_BGR2GRAY)
                            ret, thresh_out = cv2.threshold(out_Gray, 127, 255, cv2.THRESH_BINARY_INV)
                            kernel_ip = np.ones((2, 2), np.uint8)
                            eroded_ip = cv2.erode(thresh_out, kernel_ip, iterations=1)
                            dilated_ip = cv2.dilate(eroded_ip, kernel_ip, iterations=1)
                            cnts = cv2.findContours(dilated_ip.copy(), cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
                            cnts = imutils.grab_contours(cnts)
                            print(len(cnts))

                            if len(cnts) == 0:
                                    flag_empty = 1
                                    active_avi = 1
                                    flag_write = 1
                                    text = "Empty Frame"
                            else:
                                    flag_empty = 0

                            if (active_avi == 1 and flag_empty == 0):
                                    filter_out, flag_imagecaptred, flag_empty = filterImg_avi(image)  # Algorith 01

                            if flag_imagecaptred == 1:
                                    active_diff = 1
                                    # print("Difference")
                                    cropped_frame = Four_point_transform_main(filter_out)
                                    imReg = alignImages(cropped_frame, cropped_refe)  # Algorithm 02
                                    # print(imReg.shape)
                                    # print(cropped_refe.shape)
                                    images = [imReg, cropped_refe]
                                    imgs_for_diff = pad_images_to_same_size(images)

                                    cv2.waitKey(10)
                                    text, flag_imagecaptred, active_diff = diffprocessor(imgs_for_diff[0],
                                                                                         imgs_for_diff[1])
                                    print("flag is %d" % flag_write)
                                    if (flag_write == 1):
                                            csv_count += 1
                                            if text == "Falty PCB":
                                                    write_csv = "fail"
                                                    stamp.append(datetime.datetime.now().strftime("%d/%m/%Y %H:%M:%S"))
                                                    fail_counter += 1
                                            elif text == "Good PCB":
                                                    write_csv = "pass"
                                                    stamp.append(datetime.datetime.now().strftime("%d/%m/%Y %H:%M:%S"))
                                                    pass_counter += 1

                                            text_csv = "case %d" % csv_count
                                            print(text_csv)
                                            found[text_csv] = write_csv
                                            flag_write = 0

                            cv2.putText(out_img, text, (100, 100), cv2.FONT_HERSHEY_SIMPLEX, 1.27, (0, 0, 255), 2)
                            self.displayImage2(out_img, 1)
                            cv2.waitKey(10)

            writer()
            csv.close()
            cv2.destroyAllWindows()

    # ----- Object Sorting ----- #
    def Sorting_detection(self):

            cap = cv2.VideoCapture(0)
            filename = datetime.datetime.now()
            a = int(filename.strftime('%d'))
            csv = open(filename.strftime("%d %B %Y-%Hh%Mm%Ss") + ".csv", "w")
            found = {}
            prev = []
            curr = []
            breakloop = 0
            flg = 0
            flg_empty = 0

            def create_connection(db_file):
                    """ create a database connection to the SQLite database
                        specified by db_file
                    :param db_file: database file
                    :return: Connection object or None
                    """
                    conn = None
                    try:
                            conn = sqlite3.connect(db_file)
                            print('Successfull connected')
                    except Error as e:
                            print(e)
                            print('Not connected successfully')

                    return conn

            def create_task(conn, task):
                    """
                    Create a new task
                    :param conn:
                    :param task:
                    :return:
                    """

                    sql = ''' INSERT INTO sorting_function(Object,Counter)
                      VALUES(?,?) '''

                    print("SQL-line 57")
                    cur = conn.cursor()
                    print("CUR-line 61")
                    cur.execute(sql, task)
                    print("CUR-line 63")
                    return cur.lastrowid

            def writer():
                    #             stamp_count = 0
                    csv.write("{},{}\n".format("Object", "Counter"))
                    #             csv.write("{},{},{}\n".format("Object","Counter","stamp"))
                    csv.flush()
                    path_MV = os.getcwd()
                    database = path_MV + '\\db\\SKC_MV_database.db'
                    # create a database connection
                    conn = create_connection(database)

                    for key in found.keys():
                            csv.write("{},{}\n".format(key, found[key]))
                            #                 csv.write("{},{},{}\n".format(key,found[key],stamp[stamp_count]))
                            #                 stamp_count += 1
                            task_1 = (key, found[key])
                            print("TASK_1: ", task_1)
                            with conn:
                                    create_task(conn, task_1)
                                    print("create_task is complected line 79")
                            print("create_task is complected")
                            csv.flush()
                    csv.close()

            def safe_div(x, y):  # so we don't crash so often
                    if y == 0: return 0
                    return x / y

            def midpoint(ptA, ptB):
                    return ((ptA[0] + ptB[0]) * 0.5, (ptA[1] + ptB[1]) * 0.5)

            def nothing(x):  # for trackbar
                    pass

            def rescale_frame(frame, percent=80):  # make the video windows a bit smaller
                    width = int(frame.shape[1] * percent / 100)
                    height = int(frame.shape[0] * percent / 100)
                    dim = (width, height)
                    return cv2.resize(frame, dim, interpolation=cv2.INTER_AREA)

            if not cap.isOpened():
                    print("can't open camera")
                    exit()

            # showLive
            while (True & self.aa == True):
                    latest = datetime.datetime.now()
                    b = int(latest.strftime('%d'))
                    text = ""

                    # check if date is changed, if yes write the data and close the
                    # old csv and then create the new one
                    if (a != b):
                            writer()
                            csv = open(latest.strftime("%d %B %Y-%Hh%Mm%Ss") + ".csv", "w")
                            a = b

                    ret, frame = cap.read()
                    if (ret == True & self.aa == True):
                            frame = rescale_frame(frame)
                            orig = frame.copy()
                            fshape = frame.shape
                            fheight = fshape[0]
                            fwidth = fshape[1]

                            if not ret:
                                    print("cannot capture the frame")
                                    exit()
                            out_Gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
                            ret, thresh_out = cv2.threshold(out_Gray, 127, 255, cv2.THRESH_BINARY_INV)
                            cnts = cv2.findContours(thresh_out, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
                            cnts = imutils.grab_contours(cnts)
                            print(len(cnts))
                            if len(cnts) == 0:
                                    flg_empty = 1
                                    text = "frame is empty"

                            # changing slider value for Thresold,Kernal,Iteration
                            Thresh = self.Thrash_slider.value()
                            kern = self.Kernel_slider.value()
                            iter1 = self.horizontalSlider_3.value()

                            # show value of threshold,kernal and Iteration to label
                            self.Threshold_Value.setNum(Thresh)
                            self.Kernel_value.setNum(kern)
                            self.Iteration_Value.setNum(iter1)
                            ret, thresh1 = cv2.threshold(frame, Thresh, 255, cv2.THRESH_BINARY_INV)
                            kernel = np.ones((kern, kern), np.uint8)  # square image kernel used for erosion
                            dilation = cv2.dilate(thresh1, kernel, iterations=iter1)
                            erosion = cv2.erode(dilation, kernel,
                                                iterations=iter1)  # refines all edges in the binary image

                            opening = cv2.morphologyEx(erosion, cv2.MORPH_OPEN, kernel)
                            closing = cv2.morphologyEx(opening, cv2.MORPH_CLOSE, kernel)
                            closing = cv2.cvtColor(closing, cv2.COLOR_BGR2GRAY)

                            # find contours with simple approximation cv2.RETR_TREE,cv2.CHAIN_APPROX_SIMPLE
                            contours, hierarchy = cv2.findContours(closing, cv2.RETR_TREE, cv2.CHAIN_APPROX_NONE)

                            closing = cv2.cvtColor(closing, cv2.COLOR_GRAY2RGB)
                            cv2.drawContours(closing, contours, -1, (128, 255, 0), 1)

                            # focus on only the largest outline by area
                            areas = []  # list to hold all areas

                            for contour in contours:
                                    ar = cv2.contourArea(contour)
                                    areas.append(ar)

                            if len(areas) != 0:
                                    max_area = max(areas)
                                    max_area_index = areas.index(
                                            max_area)  # index of the list element with largest area
                                    cnt = contours[
                                            max_area_index]  # largest area contour is usually the viewing window itself, why?
                                    cv2.drawContours(closing, [cnt], 0, (0, 0, 255), 1)

                                    # compute the rotated bounding box of the contour
                                    box = cv2.minAreaRect(cnt)
                                    box = cv2.cv.BoxPoints(box) if imutils.is_cv2() else cv2.boxPoints(box)
                                    box = np.array(box, dtype="int")

                                    # order the points in the contour such that they appear
                                    # in top-left, top-right, bottom-right, and bottom-left
                                    # order, then draw the outline of the rotated bounding
                                    # box
                                    box = perspective.order_points(box)
                                    cv2.drawContours(orig, [box.astype("int")], -1, (0, 255, 0), 1)

                                    # loop over the original points and draw them
                                    for (x, y) in box:
                                            cv2.circle(orig, (int(x), int(y)), 5, (0, 0, 255), -1)

                                    # unpack the ordered bounding box, then compute the midpoint
                                    # between the top-left and top-right coordinates, followed by
                                    # the midpoint between bottom-left and bottom-right coordinates
                                    (tl, tr, br, bl) = box
                                    (tltrX, tltrY) = midpoint(tl, tr)
                                    (blbrX, blbrY) = midpoint(bl, br)

                                    # compute the midpoint between the top-left and top-right points,
                                    # followed by the midpoint between the top-righ and bottom-right
                                    (tlblX, tlblY) = midpoint(tl, bl)
                                    (trbrX, trbrY) = midpoint(tr, br)

                                    # draw the midpoints on the image
                                    cv2.circle(orig, (int(tltrX), int(tltrY)), 5, (255, 0, 0), -1)
                                    cv2.circle(orig, (int(blbrX), int(blbrY)), 5, (255, 0, 0), -1)
                                    cv2.circle(orig, (int(tlblX), int(tlblY)), 5, (255, 0, 0), -1)
                                    cv2.circle(orig, (int(trbrX), int(trbrY)), 5, (255, 0, 0), -1)

                                    # draw lines between the midpoints
                                    cv2.line(orig, (int(tltrX), int(tltrY)), (int(blbrX), int(blbrY)), (255, 0, 255), 1)
                                    cv2.line(orig, (int(tlblX), int(tlblY)), (int(trbrX), int(trbrY)), (255, 0, 255), 1)
                                    cv2.drawContours(orig, [cnt], 0, (0, 0, 255), 1)

                                    # compute the Euclidean distance between the midpoints
                                    dA = dist.euclidean((tltrX, tltrY), (blbrX, blbrY))
                                    dB = dist.euclidean((tlblX, tlblY), (trbrX, trbrY))

                                    # compute the size of the object
                                    pixelsPerMetric = 4  # more to do here to get actual measurements that have meaning in the real world
                                    dimA = dA / pixelsPerMetric
                                    dimB = dB / pixelsPerMetric
                                    print("dimA:", dimA)
                                    print("dimB:", dimB)

                                    # draw the object sizes on the image
                                    #                     cv2.putText(orig, "{:.1f}mm".format(dimB), (int(tltrX - 15), int(tltrY - 10)), cv2.FONT_HERSHEY_SIMPLEX, 0.65, (255, 255, 255), 2)
                                    #                     cv2.putText(orig, "{:.1f}mm".format(dimA), (int(trbrX + 10), int(trbrY)), cv2.FONT_HERSHEY_SIMPLEX, 0.65, (255, 255, 255), 2)

                                    # compute the center of the contour
                                    M = cv2.moments(cnt)
                                    cX = int(safe_div(M["m10"], M["m00"]))
                                    cY = int(safe_div(M["m01"], M["m00"]))

                                    # draw the contour and center of the shape on the image
                                    cv2.circle(orig, (cX, cY), 5, (0, 0, 0), -1)
                                    #                     cv2.putText(orig, "center", (cX - 20, cY - 20), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 0, 0), 2)
                                    dim_A = 46.4
                                    dim_B = 75.9
                                    ###sharpner dimentions
                                    dimsharp_A = 20
                                    dimsharp_B = 30
                                    ###quality bedge dimentions
                                    dimqual_A = 58
                                    dimqual_B = 58

                                    if ((((dim_A + 3) >= dimA >= (dim_A - 3)) & (dim_B + 3 >= dimB >= (dim_B - 3))) or (
                                            ((dim_A + 3) >= dimB >= (dim_A - 3)) & (dim_B + 3 >= dimA >= (dim_B - 3)))):
                                            # print("This is PCB-Image")
                                            flg = 1
                                            text = "PCB Image"
                                            # cv2.imshow(windowName, orig)
                                    if ((((dimsharp_A + 1) >= dimA >= (dimsharp_A - 2)) & (
                                            dimsharp_B + 1 >= dimB >= (dimsharp_B - 2))) or (
                                            ((dimsharp_A + 1) >= dimB >= (dimsharp_A - 2)) & (
                                            dimsharp_B + 1 >= dimA >= (dimsharp_B - 2)))):
                                            # print("This is sharpner")
                                            flg = 1
                                            text = "Sharpner"
                                            # cv2.putText(orig, text, (cX - 50, cY - 50), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 0, 255), 2)
                                            # cv2.imshow(windowName, orig)
                                    if ((((dimqual_A + 3) >= dimA >= (dimqual_A - 3)) & (
                                            dimqual_B + 3 >= dimB >= (dimqual_B - 3))) or (
                                            ((dimqual_A + 3) >= dimB >= (dimqual_A - 3)) & (
                                            dimqual_B + 3 >= dimA >= (dimqual_A - 3)))):
                                            # print("This is bedge")
                                            flg = 1
                                            text = "Quality Badge"

                                    print(flg, flg_empty, text)
                                    if (flg == 1 and flg_empty == 1):
                                            flg = 0
                                            flg_empty = 0
                                            if text not in found.keys():
                                                    found[text] = 1

                                            else:
                                                    found[text] = found[text] + 1

                            cv2.putText(orig, text, (50, 50), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 0, 255), 2)
                            self.displayImage2(orig, 1)
                            cv2.waitKey(10)
                            flg = 0

            writer()
            csv.close()
            cap.release()
            cv2.destroyAllWindows()

    # ----- Object Dimension ----- #
    def Object_Dimensio(self):
            # using cam built-in to computer
            #         In_Cam= self.Input_Camera()
            #         cap = cv2.VideoCapture(In_Cam
            cap = cv2.VideoCapture(0)
            start_time=time.time()
            def safe_div(x, y):  # so we don't crash so often
                    if y == 0: return 0
                    return x / y

            def nothing(x):  # for trackbar
                    pass

            def rescale_frame(frame, percent=80):  # make the video windows a bit smaller
                    width = int(frame.shape[1] * percent / 100)
                    height = int(frame.shape[0] * percent / 100)
                    dim = (width, height)
                    return cv2.resize(frame, dim, interpolation=cv2.INTER_AREA)

            if not cap.isOpened():
                    print("can't open camera")
                    exit()

            showLive = True
            while ((cap.isOpened() & self.aa == True)):

                    ret, frame = cap.read()
                    if not ret:
                            print("cannot capture the frame")
                            exit()
                    frame = rescale_frame(frame)
                    out_Gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
                    ret, thresh_out = cv2.threshold(out_Gray, 127, 255, cv2.THRESH_BINARY_INV)
                    cnts = cv2.findContours(thresh_out, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
                    cnts = imutils.grab_contours(cnts)

                    if len(cnts) == 0:
                            text = "frame is empty"
                            cv2.putText(frame, text, (25, 25), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 0, 255), 2)
                            self.displayImage2(frame, 1)
                            cv2.waitKey(10)
                            continue

                    # changing slider value for Thresold,Kernal,Iteration
                    Thresh = self.Thrash_slider.value()
                    kern = self.Kernel_slider.value()
                    iter1 = self.horizontalSlider_3.value()

                    # show value of threshold,kernal and Iteration to label
                    self.Threshold_Value.setNum(Thresh)
                    self.Kernel_value.setNum(kern)
                    self.Iteration_Value.setNum(iter1)

                    ret, thresh1 = cv2.threshold(frame, Thresh, 255, cv2.THRESH_BINARY_INV)
                    kernel = np.ones((kern, kern), np.uint8)  # square image kernel used for erosion
                    dilation = cv2.dilate(thresh1, kernel, iterations=iter1)
                    erosion = cv2.erode(dilation, kernel, iterations=iter1)  # refines all edges in the binary image
                    opening = cv2.morphologyEx(erosion, cv2.MORPH_OPEN, kernel)
                    closing = cv2.morphologyEx(opening, cv2.MORPH_CLOSE, kernel)
                    closing = cv2.cvtColor(closing, cv2.COLOR_BGR2GRAY)

                    #             ret,contours,hierarchy = cv2.findContours(closing,cv2.RETR_TREE,cv2.CHAIN_APPROX_NONE) # find contours with simple approximation cv2.RETR_TREE,cv2.CHAIN_APPROX_SIMPLE
                    contours, hierarchy = cv2.findContours(closing, cv2.RETR_TREE, cv2.CHAIN_APPROX_NONE)
                    closing = cv2.cvtColor(closing, cv2.COLOR_GRAY2RGB)
                    cv2.drawContours(closing, contours, -1, (128, 255, 0), 1)

                    # focus on only the largest outline by area
                    areas = []  # list to hold all areas

                    for contour in contours:
                            ar = cv2.contourArea(contour)
                            areas.append(ar)
                    if len(areas) == 0:
                            continue

                    max_area = max(areas)
                    max_area_index = areas.index(max_area)  # index of the list element with largest area
                    cnt = contours[max_area_index]  # largest area contour is usually the viewing window itself, why?
                    cv2.drawContours(closing, [cnt], 0, (0, 0, 255), 1)

                    def midpoint(ptA, ptB):
                            return ((ptA[0] + ptB[0]) * 0.5, (ptA[1] + ptB[1]) * 0.5)

                    # compute the rotated bounding box of the contour
                    orig = frame.copy()
                    box = cv2.minAreaRect(cnt)
                    box = cv2.cv.BoxPoints(box) if imutils.is_cv2() else cv2.boxPoints(box)
                    box = np.array(box, dtype="int")

                    # order the points in the contour such that they appear
                    # in top-left, top-right, bottom-right, and bottom-left
                    # order, then draw the outline of the rotated bounding
                    # box
                    box = perspective.order_points(box)
                    cv2.drawContours(orig, [box.astype("int")], -1, (0, 255, 0), 1)

                    # loop over the original points and draw them
                    for (x, y) in box:
                            cv2.circle(orig, (int(x), int(y)), 5, (0, 0, 255), -1)

                    # unpack the ordered bounding box, then compute the midpoint
                    # between the top-left and top-right coordinates, followed by
                    # the midpoint between bottom-left and bottom-right coordinates
                    (tl, tr, br, bl) = box
                    (tltrX, tltrY) = midpoint(tl, tr)
                    (blbrX, blbrY) = midpoint(bl, br)

                    # compute the midpoint between the top-left and top-right points,
                    # followed by the midpoint between the top-righ and bottom-right
                    (tlblX, tlblY) = midpoint(tl, bl)
                    (trbrX, trbrY) = midpoint(tr, br)

                    # draw the midpoints on the image
                    cv2.circle(orig, (int(tltrX), int(tltrY)), 5, (255, 0, 0), -1)
                    cv2.circle(orig, (int(blbrX), int(blbrY)), 5, (255, 0, 0), -1)
                    cv2.circle(orig, (int(tlblX), int(tlblY)), 5, (255, 0, 0), -1)
                    cv2.circle(orig, (int(trbrX), int(trbrY)), 5, (255, 0, 0), -1)

                    # draw lines between the midpoints
                    cv2.line(orig, (int(tltrX), int(tltrY)), (int(blbrX), int(blbrY)), (255, 0, 255), 1)
                    cv2.line(orig, (int(tlblX), int(tlblY)), (int(trbrX), int(trbrY)), (255, 0, 255), 1)
                    cv2.drawContours(orig, [cnt], 0, (0, 0, 255), 1)

                    # compute the Euclidean distance between the midpoints
                    dA = dist.euclidean((tltrX, tltrY), (blbrX, blbrY))
                    dB = dist.euclidean((tlblX, tlblY), (trbrX, trbrY))

                    # compute the size of the object
                    pixelsPerMetric = 4  # more to do here to get actual measurements that have meaning in the real world
                    dimA = dA / pixelsPerMetric
                    dimB = dB / pixelsPerMetric
                    # SKC#

                    # SKC
                    # draw the object sizes on the image
                    cv2.putText(orig, "{:.1f}mm".format(dimB), (int(tltrX - 15), int(tltrY - 10)),
                                cv2.FONT_HERSHEY_SIMPLEX, 0.65, (0, 0, 0), 2)
                    cv2.putText(orig, "{:.1f}mm".format(dimA), (int(trbrX + 10), int(trbrY)), cv2.FONT_HERSHEY_SIMPLEX,
                                0.65, (0, 0, 0), 2)

                    # compute the center of the contour
                    M = cv2.moments(cnt)
                    cX = int(safe_div(M["m10"], M["m00"]))
                    cY = int(safe_div(M["m01"], M["m00"]))

                    # draw the contour and center of the shape on the image
                    cv2.circle(orig, (cX, cY), 5, (255, 255, 255), -1)
                    print("-----%s seconds---" % (time.time()-start_time))
                    #             cv2.putText(orig, "center", (cX - 20, cY - 20), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (255, 0, 255), 2)
                    self.displayImage2(orig, 1)
                    if cv2.waitKey(30) >= 0:
                            showLive = False
            
            cap.release()
            cv2.destroyAllWindows()

    # Selection of QR / Barcode
    def Scan_QR_Barcode(self):
            total = "Kalyan"
            print(total)
            self.OutPutData.setText(total)
            # initialize the video stream and allow the camera sensor to warm up
            #         In_Cam= self.Input_Camera()
            #         cap = cv2.VideoCapture(In_Cam)
            cap = cv2.VideoCapture(0)
            print("[INFO] starting video stream...")
            time.sleep(2.0)

            # open the output CSV file for writing and initialize the set of
            # barcodes found thus far and take date at the time of running the code
            filename = datetime.datetime.now()
            a = int(filename.strftime('%d'))
            csv = open(filename.strftime("%d %B %Y-%Hh%Mm%Ss") + ".csv", "w")
            found = {}
            prev = []
            curr = []
            z=[]

            def writer():
                    csv.write("{},{}\n".format("barcode data", "counter"))
                    csv.flush()

                    for key in found.keys():
                            csv.write("{},{}\n".format(key, found[key]))
                            csv.flush()
                    csv.close()

            def rescale_frame(frame, percent=80):  # make the video windows a bit smaller
                    width = int(frame.shape[1] * percent / 100)
                    height = int(frame.shape[0] * percent / 100)
                    dim = (width, height)
                    return cv2.resize(frame, dim, interpolation=cv2.INTER_AREA)

            # loop over the frames from the video stream
            while (True & self.aa == True):
                    # grab the frame from the threaded video stream and resize it to
                    # have a maximum width of 400 pixels
                    latest = datetime.datetime.now()
                    b = int(latest.strftime('%d'))

                    # check if date is changed, if yes write the data and close the
                    # old csv and then create the new one
                    if (a != b):
                            writer()
                            csv = open(latest.strftime("%d %B %Y-%Hh%Mm%Ss") + ".csv", "w")
                            a = b

                    ret, frame = cap.read()
                    # frame = imutils.resize(frame, width=400)

                    if (ret == True & self.aa == True):
                            # find the barcodes in the frame and decode each of the barcodes
                            frame = rescale_frame(frame)
                            out_new = np.uint8(frame)
                            out_Gray = cv2.cvtColor(out_new, cv2.COLOR_BGR2GRAY)
                            ret, thresh_out = cv2.threshold(out_Gray, 127, 255, cv2.THRESH_BINARY_INV)
                            kernel_ip = np.ones((2, 2), np.uint8)
                            eroded_ip = cv2.erode(thresh_out, kernel_ip, iterations=1)
                            dilated_ip = cv2.dilate(eroded_ip, kernel_ip, iterations=1)
                            cnts = cv2.findContours(dilated_ip.copy(), cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
                            cnts = imutils.grab_contours(cnts)
                            #print(len(cnts))

                            if len(cnts) == 0:
                                    flg = 0
                                    text = 'waiting for barcode'
                                    cv2.putText(frame, text, (25, 25), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 0, 255), 2)
                                    self.displayImage2(frame, 1)
                                    cv2.waitKey(10)
                                    continue

                            barcodes = pyzbar.decode(frame)
                            #                 print(barcodes)
                            #                 if barcodes == []:
                            #                     flg = 0
                            
                            # loop over the detected barcodes
                            for barcode in barcodes:
                                    # extract the bounding box location of the barcode and draw
                                    # the bounding box surrounding the barcode on the image
                                    (x, y, w, h) = barcode.rect
                                    cv2.rectangle(frame, (x, y), (x + w, y + h), (0, 0, 255), 2)

                                    # the barcode data is a bytes object so if we want to draw it
                                    # on our output image we need to convert it to a string first
                                    barcodeData = barcode.data.decode("utf-8")
                                    prev.append(barcodeData)
                                    barcodeType = barcode.type

                                    # draw the barcode data and barcode type on the image
                                    text = "{} ({})".format(barcodeData, barcodeType)
                                    cv2.putText(frame, text, (x, y - 10),
                                                cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 0, 255), 2)
                                                
                                    # ----- total = "Kalyan" ----- #
                                    self.OutPutData.setText(barcodeData)
                                    aa=str(barcodeData)
                                    z.append(barcodeData)
                                    windowstring = ''
                                    nline = '\n'
                                    for i in z:
                                        windowstring =  windowstring+i+nline
                                    self.AllOutPutData.setText(windowstring)
                                    # ----- total = "Kalyan" ----- #
                                    
                                    flg = 0
                                    # if the barcode text is currently not in our CSV file, write
                                    # the timestamp + barcode to disk and update the set
                                    if flg == 0:
                                            if barcodeData not in found.keys():
                                                    found[barcodeData] = 1
                                            else:
                                                    found[barcodeData] = found[barcodeData] + 1
                                            flg = 1

                            # show the output frame
                            # cv2.imshow("Barcode Scanner", frame)
                            self.displayImage2(frame, 1)
                            time.sleep(0.05)

            # close the output CSV file do a bit of cleanup
            #print("[INFO] cleaning up...")
            writer()
            csv.close()
            cap.release()
            cv2.destroyAllWindows()

    def StartOption(self):
            print("Start Option")
            self.completed = 0

            while self.completed < 100:
                    self.completed += 0.0001
                    self.progressBar.setValue(self.completed)

    # convert frame to image format to show in label
    def displayImage2(self, img1, window=1):
            qformat = QImage.Format_Indexed8
            if len(img1.shape) == 3:

                    if (img1.shape[2]) == 4:
                            qformat = QImage.Format_RGBA8888
                    else:
                            qformat = QImage.Format_RGB888
            img1 = QImage(img1, img1.shape[1], img1.shape[0], qformat)
            img1 = img1.rgbSwapped()
            QApplication.processEvents()
            self.OutPutScreen.setPixmap(QPixmap.fromImage(img1))
            return img1

    # ----- Exit Function ------#
    def Exit_Screen(self):
            print("Successfully Exited \n")
            sys.exit(app.exec_())
            exit()

    # ----- Reset Function ------#
    def ResetOption(self):
            print("Reset Successfully")
            self.aa = False
            reset_flg = 1
            cv2.destroyAllWindows()
            return reset_flg

    # calling main window and passing parameter for Respective color
    def ColorWindow(self):
            print("Enter Color Window  Def")
            self.dialog = Ui_second_Window()
            self.dialog.submitted.connect(self.Color_Identification)
            self.dialog.show()
            print("kalyan")



if __name__ == "__main__":
    import sys
    app = QtWidgets.QApplication(sys.argv)
    MainWindow = QtWidgets.QMainWindow()
    ui = Ui_MainWindow()
    ui.setupUi(MainWindow)
    MainWindow.show()
    sys.exit(app.exec_())
